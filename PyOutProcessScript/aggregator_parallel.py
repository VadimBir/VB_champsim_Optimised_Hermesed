#!/usr/bin/env python3
"""ChampSim results aggregator — PARALLEL variant (aggregator_parallel.py).

Identical to aggregator.py except collect() uses ProcessPoolExecutor to
parse log files in parallel.  Default: 32 workers, chunksize=256.
Override via --workers N.

Usage:
    python3 aggregator_parallel.py <outputs-dir> [--workers N]

See plan: ~/.claude/plans/cryptic-plotting-noodle.md
"""

import atexit
import os
import re
import signal
import subprocess
import sys
import time
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor

# CP9: global failure tracker — appended during collect(), surfaced by
# _print_failure_report() via atexit + SIGINT/SIGTERM trap so even an
# interrupted run prints the list of bad logs before exiting.
FAILED_LOGS = []  # list of (path, reason)


def _print_failure_report():
    if not FAILED_LOGS:
        return
    print("=== WARNING: failed/incomplete logs ===", file=sys.stderr)
    for p, reason in FAILED_LOGS[:200]:
        print(f"  {os.path.basename(p)}  reason: {reason}", file=sys.stderr)
    if len(FAILED_LOGS) > 200:
        print(f"  ... and {len(FAILED_LOGS) - 200} more", file=sys.stderr)
    print(f"  total: {len(FAILED_LOGS)} failed/incomplete", file=sys.stderr)


def _signal_trap(signum, _frame):
    _print_failure_report()
    sys.exit(128 + signum)


atexit.register(_print_failure_report)
signal.signal(signal.SIGINT, _signal_trap)
signal.signal(signal.SIGTERM, _signal_trap)

from openpyxl import Workbook

from stat_patterns import (
    STAT_PATTERNS,
    ERROR_MARKERS,
    PER_LEVEL_FIELD_RE,
    SUM_FIELDS,
    AVG_FIELDS,
    CP5_FAMILIES,
    LEVELS_TRIPLE,
)

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import ColorScaleRule
# ColorScaleRule removed — no conditional formatting in simplified version

PER_LEVEL_RE_C = re.compile(PER_LEVEL_FIELD_RE)

# cd to script dir so relative paths work
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(SCRIPT_DIR)

OUTPUT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "999_READY_EXCELS"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Filename format (v14-9, double-dot separator):
#   <ARCH>..<N>c..<TRACE>..<L1pf>..<L2pf>..<L3pf>..<BP>..<REPL>..ByP..<L1m>..<L2m>..<L3m>..<WIN_MODE>..<HERMES>.log
# Separator is `..` (double dot) because single `.` collides with trace names (e.g. 437.leslie3d).
# TRACE may contain single `.` and `-` (e.g. LLM1024.GPTpre30Phase15xpost20-496M_208M, 437.leslie3d-273B).
# Models may contain `-` (e.g. 4000fix-KappaPhiL1L2). HERMES ∈ {off, hermes, ttp, hmp}.
# Parsed via split('..') since `..` never appears inside any field.
FILENAME_RE = None  # legacy; new parser uses split-based logic in parse_log_filename

CORE_DIR_RE = re.compile(r"^core(\d+)$")

SENTINEL_RUNNING = -1
SENTINEL_ERRORED = -2

# ── FILTER ARRAYS (edit these to control which traces/models appear) ─────
# Master switch: when False, ALLOWED_TRACES/ALLOWED_MODELS are ignored
# (all logs included). Prints a warning at startup if disabled.
FILTER_INPUT = False
ALLOWED_TRACES = {
    "403.gcc-17B", "403.gcc-48B", "410.bwaves-1963B", "410.bwaves-2097B",
    "429.mcf-184B", "429.mcf-192B", "429.mcf-22B", "429.mcf-51B",
    "433.milc-127B", "433.milc-274B", "433.milc-337B",
    "437.leslie3d-134B", "437.leslie3d-149B", "437.leslie3d-232B",
    "437.leslie3d-265B", "437.leslie3d-271B", "437.leslie3d-273B",
    "450.soplex-247B", "450.soplex-92B",
    "459.GemsFDTD-1211B", "459.GemsFDTD-1320B", "459.GemsFDTD-1418B",
    "459.GemsFDTD-1491B", "459.GemsFDTD-765B",
    "462.libquantum-1343B", "462.libquantum-714B",
    "470.lbm-1274B", "471.omnetpp-188B", "473.astar-359B",
    "481.wrf-1254B", "481.wrf-196B", "481.wrf-816B",
    "482.sphinx3-234B",
    "483.xalancbmk-127B", "483.xalancbmk-716B", "483.xalancbmk-736B",
    "602.gcc_s-1850B", "602.gcc_s-734B",
    "603.bwaves_s-1740B", "603.bwaves_s-2931B",
    "605.mcf_s-1152B", "605.mcf_s-1554B", "605.mcf_s-1644B",
    "605.mcf_s-484B", "605.mcf_s-665B", "605.mcf_s-782B",
    "607.cactuBSSN_s-3477B", "607.cactuBSSN_s-4004B",
    "619.lbm_s-2677B", "619.lbm_s-4268B",
    "620.omnetpp_s-141B", "620.omnetpp_s-874B",
    "621.wrf_s-8100B",
    "623.xalancbmk_s-10B", "623.xalancbmk_s-202B",
    "627.cam4_s-490B",
    "649.fotonik3d_s-1176B", "649.fotonik3d_s-7084B", "649.fotonik3d_s-8225B",
    "654.roms_s-1070B", "654.roms_s-1390B", "654.roms_s-294B", "654.roms_s-523B",
    "LLM1024.GPT-125M_496M", "LLM1024.Pythia-70M_331M",
    "LLM256.GPT-125M_32M", "LLM256.OPT-350M_42M", "LLM256.Pythia-70M_21M",
    "LLM512.GPT-125M_125M", "LLM512.OPT-350M_167M", "LLM512.Pythia-70M_83M",
}

# No model whitelist — all models are accepted.
ALLOWED_MODELS = set()


def split_model_triplet(modelstr):
    """Model segment is <L1m>-<L2m>-<LLCm> (v14: hyphen-joined). Models may
    themselves contain hyphens (e.g. 4000-KappaPhiL1L2-...).

    Strategy: try equal-thirds split on hyphen boundary; require homogeneous
    (a==b==c). Fallback: whole string as single-model triplet.
    """
    parts = modelstr.split("-")
    n = len(parts)
    if n % 3 == 0:
        third = n // 3
        a = "-".join(parts[:third])
        b = "-".join(parts[third : 2 * third])
        c = "-".join(parts[2 * third :])
        if a == b == c:
            return a, b, c
    return modelstr, modelstr, modelstr


_NUM_PREFIX_RE = re.compile(r"^\d+-")


def parse_log_filename(fname):
    # Accept both <name>.log and <name>.log.INV_DDL_FIX (invisible-deadlock backup).
    if fname.endswith(".log.INV_DDL_FIX"):
        base = fname[: -len(".log.INV_DDL_FIX")]
    elif fname.endswith(".log"):
        base = fname[:-4]
    else:
        return None
    # Strip leading numeric prefix like "000-" (e.g. 000-glc..16c.. == glc..16c..).
    base = _NUM_PREFIX_RE.sub("", base)
    parts = base.split("..")
    if len(parts) != 14:
        return None
    (arch, cores_s, trace, l1pf, l2pf, l3pf, bp_pred, repl,
     byp_lit, l1m, l2m, l3m, win_mode, hermes) = parts
    if byp_lit != "ByP":
        return None
    if not cores_s.endswith("c"):
        return None
    try:
        cores = int(cores_s[:-1])
    except ValueError:
        return None
    if hermes not in ("off", "hermes", "ttp", "hmp"):
        return None
    return {
        "bp": arch,
        "l1pf": l1pf,
        "l2pf": l2pf,
        "llcpf": l3pf,
        "repl": repl,
        "cores": cores,
        "win_mode": win_mode,
        "hermes": hermes,
        "trace": trace,
        "l1m": l1m,
        "l2m": l2m,
        "llcm": l3m,
        "branch": bp_pred,
    }


def read_log(path):
    try:
        with open(path, "r", errors="replace") as f:
            return f.read()
    except OSError:
        return None


ERROR_MARKER_RE = re.compile(
    r"(?:^|[\s:/])(?:assert(?:ion)?\s+failed|Segmentation\s+fault|SIGSEGV|"
    r"Aborted|Killed|deadlock)(?:[\s:]|$)",
    re.MULTILINE,
)


def has_error_marker(text):
    # RULE 0: avoid substring false positives like 'assertion_rate'.
    # Require anchored markers with boundary characters.
    return bool(ERROR_MARKER_RE.search(text))


ARCH_RE = re.compile(r"\[RUN\]\s+arch\s+base:\s*(\S+)")
# v14: '[BUILD] cached binary: .../bin_cache/<arch>_<cores>_..../champsim'
BUILD_BIN_ARCH_RE = re.compile(r"\[BUILD\][^\n]*?bin_cache/([^/_]+)_\d+_")


def extract_arch(text):
    # Prefer v14 build-binary path, fall back to legacy [RUN] arch base.
    if text is not None:
        m = BUILD_BIN_ARCH_RE.search(text)
        if m:
            return m.group(1)
        m = ARCH_RE.search(text)
        if m:
            return m.group(1)
    return None


INV_DDL_IPC_RE = re.compile(r"INVISIBLE DEADLOCK\s+\d+/\d+\s+AVG IPC:\s*([\d.]+)")


def extract_ipc(text):
    """Return (value, sentinel_used). sentinel_used in {0,-1,-2}.

    Tries FINAL ROI first; falls back to invisible-deadlock-fix AVG IPC line
    (appended by scripts/fix_invisible_deadlock.sh into .log.INV_DDL_FIX files).
    """
    if text is None:
        return SENTINEL_ERRORED, -2
    regex, _scope = STAT_PATTERNS["ipc_final_avg"]
    m = re.search(regex, text)
    if m:
        try:
            return float(m.group(1)), 0
        except ValueError:
            pass
    m = INV_DDL_IPC_RE.search(text)
    if m:
        try:
            return float(m.group(1)), 0
        except ValueError:
            pass
    # no value found — classify
    if has_error_marker(text):
        return SENTINEL_ERRORED, -2
    return SENTINEL_RUNNING, -1


def extract_per_level_fields(text):
    """Parse all `Core_;N; LVL;_;field; val;` lines.

    Returns {(lvl, field): aggregated_scalar}. Aggregation:
      - SUM_FIELDS → sum across cores
      - AVG_FIELDS → average across cores (skipping nan/inf)
      - else → average (default)
    Missing field → not in dict (caller substitutes SENTINEL_RUNNING).
    """
    if text is None:
        return {}
    # (lvl, field) -> list of per-core floats
    raw = defaultdict(list)
    for m in PER_LEVEL_RE_C.finditer(text):
        _core, lvl, field, val_s = m.group(1), m.group(2), m.group(3), m.group(4)
        if val_s in ("nan", "inf", "-nan", "-inf"):
            continue
        try:
            v = float(val_s)
        except ValueError:
            continue
        raw[(lvl, field)].append(v)
    out = {}
    for (lvl, field), vals in raw.items():
        if not vals:
            continue
        if field in SUM_FIELDS:
            out[(lvl, field)] = sum(vals)
        elif field in AVG_FIELDS:
            out[(lvl, field)] = sum(vals) / len(vals)
        else:
            out[(lvl, field)] = sum(vals) / len(vals)
    return out


LPM_HEADER_RE = re.compile(r"=== LPM Cycle Classification\s+CPU\s+(\d+)\s+\(ROI\) ===")
LPM_COLS = (
    "h", "m", "x", "e",
    "omega", "mu", "kappa", "phi",
    "LPMR_std", "LPMR_byp",
    "C-AMAT", "APC", "MST", "CPA",
)
LPM_LEVELS = ("ITLB", "DTLB", "STLB", "L1I", "L1D", "L2C", "LLC", "DRAM")


def extract_lpm_cycle_class(text):
    """Parse all `=== LPM Cycle Classification CPU N (ROI) ===` blocks.

    Returns {(lvl, 'LPM_'+col): aggregated_scalar} averaged across cores.
    Values of 'n/a' / 'nan' / 'inf' are skipped.
    """
    if text is None:
        return {}
    raw = defaultdict(list)
    lines = text.splitlines()
    i = 0
    n = len(lines)
    while i < n:
        if LPM_HEADER_RE.search(lines[i]):
            # skip header + column-label line
            i += 2
            while i < n:
                line = lines[i].strip()
                if not line:
                    break
                # strip | separators
                parts = re.split(r"\s*\|\s*|\s+", line)
                parts = [p for p in parts if p]
                if not parts or parts[0] not in LPM_LEVELS:
                    break
                lvl = parts[0]
                vals = parts[1:]
                # Expect 14 columns. Tolerate short rows.
                for c_idx, col_name in enumerate(LPM_COLS):
                    if c_idx >= len(vals):
                        break
                    v_s = vals[c_idx]
                    if v_s in ("n/a", "nan", "-nan", "inf", "-inf"):
                        continue
                    try:
                        v = float(v_s)
                    except ValueError:
                        continue
                    raw[(lvl, "LPM_" + col_name)].append(v)
                i += 1
            continue
        i += 1
    out = {}
    for k, vals in raw.items():
        if vals:
            out[k] = sum(vals) / len(vals)
    return out


BLT_HEADER_RE = re.compile(r"=== ByPLatTracker CPU\s+(\d+)\s+===")
BLT_LEVELS = {"L1D", "L2C", "LLC"}


def extract_byplat_tracker(text):
    """Parse ByPLatTracker blocks. Supports two header formats:
    Old: Lvl mshr_avg mshr_pk byp_infl_avg byp_pk mlp_avg mlp_pk samples
    New: Lvl lat_global lat_short mshr_avg byp_avg mlp_avg mlp_pk
    Returns {(lvl, 'BLT_<field>'): avg_across_cores}.
    """
    if text is None:
        return {}
    raw = defaultdict(list)
    lines = text.splitlines()
    i = 0
    n = len(lines)
    while i < n:
        if BLT_HEADER_RE.search(lines[i]):
            i += 1
            if i >= n:
                break
            hdr_parts = lines[i].split()
            if not hdr_parts or hdr_parts[0] != "Lvl":
                i += 1
                continue
            col_names = hdr_parts[1:]
            i += 1
            while i < n:
                parts = lines[i].split()
                if not parts or parts[0] not in BLT_LEVELS:
                    break
                lvl = parts[0]
                vals = parts[1:]
                col_map = {}
                for ci, cname in enumerate(col_names):
                    if ci >= len(vals):
                        break
                    col_map[cname] = vals[ci]
                mshr = col_map.get("mshr_avg")
                mshr_pk = col_map.get("mshr_pk")
                byp = col_map.get("byp_infl_avg") or col_map.get("byp_avg")
                mlp = col_map.get("mlp_avg")
                mlp_pk = col_map.get("mlp_pk")
                for fname, val_s in [("BLT_mshr_avg", mshr), ("BLT_mshr_pk", mshr_pk),
                                     ("BLT_byp_avg", byp),
                                     ("BLT_mlp_avg", mlp), ("BLT_mlp_pk", mlp_pk)]:
                    if val_s is None:
                        continue
                    try:
                        v = float(val_s)
                    except ValueError:
                        continue
                    raw[(lvl, fname)].append(v)
                i += 1
            continue
        i += 1
    out = {}
    for k, vals in raw.items():
        if vals:
            out[k] = sum(vals) / len(vals)
    return out


def _process_log(args):
    """Worker function — runs in a separate process.

    args: (fpath, fname, cores, dir_base)
    Returns: (key_or_None, metrics_or_None, sentinel, fpath, fail_reason_or_None)
      sentinel: 0=ok  -1=running  -2=errored  None=unparsed
    """
    fpath, fname, cores, dir_base = args
    meta = parse_log_filename(fname)
    if meta is None:
        return None, None, None, fpath, "unparsed"
    # Apply filters early — skip before reading log content
    if FILTER_INPUT and ALLOWED_TRACES and meta["trace"] not in ALLOWED_TRACES:
        return None, None, None, fpath, None
    model_combo = f'{meta["l1m"]}+{meta["l2m"]}+{meta["llcm"]}'
    if FILTER_INPUT and ALLOWED_MODELS and model_combo not in ALLOWED_MODELS:
        return None, None, None, fpath, None
    text = read_log(fpath)
    arch = extract_arch(text) or meta.get("bp") or "glc"
    ipc, sentinel = extract_ipc(text)

    is_inv_ddl = fname.endswith(".log.INV_DDL_FIX")

    fail_reason = None
    if sentinel == -1:
        fail_reason = "incomplete (no FINAL ROI, no error marker)"
    elif sentinel == -2:
        fail_reason = "error marker found"
        for em in ERROR_MARKERS:
            if re.search(rf"\b{re.escape(em)}\b", text or ""):
                fail_reason = f"error: {em}"
                break

    # For .INV_DDL_FIX (invisible-deadlock backup) skip per-level — partial-run
    # numbers are unreliable; user only wants the appended AVG IPC.
    if is_inv_ddl:
        per_level = {}
    else:
        per_level = extract_per_level_fields(text)
        per_level.update(extract_lpm_cycle_class(text))
        per_level.update(extract_byplat_tracker(text))

    key = (
        arch,
        cores,
        meta["trace"],
        meta["l1pf"],
        meta["l2pf"],
        meta["llcpf"],
        meta["l1m"],
        meta["l2m"],
        meta["llcm"],
        meta["repl"],
        meta["hermes"],
    )
    metrics = {"ipc": ipc}
    metrics.update(per_level)
    return key, metrics, sentinel, fpath, fail_reason


def collect(outputs_dir, n_workers=32, chunksize=256):
    outputs_dir = os.path.abspath(outputs_dir)
    dir_base = os.path.basename(outputs_dir.rstrip("/"))

    # Phase 1: enumerate all tasks
    tasks = []
    unparsed_names = []
    counts = {"logs": 0, "parsed": 0, "ok": 0, "running": 0, "errored": 0}

    for entry in sorted(os.listdir(outputs_dir)):
        core_match = CORE_DIR_RE.match(entry)
        if not core_match:
            continue
        cores = int(core_match.group(1))
        core_dir = os.path.join(outputs_dir, entry)
        if not os.path.isdir(core_dir):
            continue
        # Enumerate .log first (real sims), then .log.INV_DDL_FIX (deadlock backups).
        # Order matters: collect() merge rule prefers earlier non-sentinel entries.
        entries = sorted(os.listdir(core_dir))
        log_files = [f for f in entries if f.endswith(".log")]
        inv_files = [f for f in entries if f.endswith(".log.INV_DDL_FIX")]
        for fname in log_files + inv_files:
            counts["logs"] += 1
            tasks.append((os.path.join(core_dir, fname), fname, cores, dir_base))

    # Phase 2: parallel parse
    data = {}
    total = len(tasks)
    done = 0
    PROGRESS_INTERVAL = 500
    next_report = PROGRESS_INTERVAL
    t0 = time.time()
    with ProcessPoolExecutor(max_workers=n_workers) as executor:
        for key, metrics, sentinel, fpath, fail_reason in executor.map(
            _process_log, tasks, chunksize=chunksize
        ):
            done += 1
            if sentinel is None:
                unparsed_names.append(os.path.basename(fpath))
            else:
                counts["parsed"] += 1
                if sentinel == 0:
                    counts["ok"] += 1
                elif sentinel == -1:
                    counts["running"] += 1
                    FAILED_LOGS.append((fpath, fail_reason))
                else:
                    counts["errored"] += 1
                    FAILED_LOGS.append((fpath, fail_reason))
                # Merge rule: prefer entry with a real IPC (sentinel >= 0).
                # Order is .log first, .INV_DDL_FIX second, so complete .log
                # naturally wins; an INV_DDL_FIX only fills in when the .log
                # was running/incomplete.
                existing = data.get(key)
                if existing is None:
                    data[key] = metrics
                elif existing.get("ipc", SENTINEL_RUNNING) < 0 and metrics.get("ipc", SENTINEL_RUNNING) >= 0:
                    # Preserve per-level data from existing if new is INV_DDL-only.
                    merged = dict(existing)
                    merged.update(metrics)
                    data[key] = merged
            if done >= next_report:
                elapsed = time.time() - t0
                rate = done / elapsed if elapsed > 0 else 0
                eta = (total - done) / rate if rate > 0 else 0
                print(f"{done}/{total} ok={counts['ok']} e={counts['errored']} ~{eta:.0f}s", flush=True)
                next_report += PROGRESS_INTERVAL

    return data, counts, unparsed_names


def deduce_missing(data):
    """CP8 test-matrix deduction.

    Build the expected matrix = cross-product of observed dimensions, then
    insert SENTINEL_RUNNING placeholder entries for any expected key that
    has no log at all. Keys with a present-but-errored log stay as-is (they
    already carry SENTINEL_ERRORED in their metric values).

    Expected rule (conservative, per-cores-count scoping):
      For each (arch, cores) pair, take the cross-product of:
        traces observed at that (arch, cores)
      × config-tuples (l1pf, l2pf, llcpf, l1m, l2m, llcm) observed at
        that (arch, cores).
      If a resulting key is not in data, insert it with a placeholder
      metrics dict (ipc = -1, no per-level fields). Downstream sheet
      builders already substitute SENTINEL_RUNNING for missing metric keys,
      so inserting just {"ipc": -1} is enough to surface the row.

    Scoping per (arch, cores) rather than globally avoids exploding the
    matrix across unrelated core counts (e.g. 1c runs a different suite
    than 16c). Returns (n_added, n_total_expected).
    """
    from collections import defaultdict as _dd
    per_ac_traces = _dd(set)
    per_ac_configs = _dd(set)

    for key in data.keys():
        arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes = key
        per_ac_traces[(arch, cores)].add(trace)
        per_ac_configs[(arch, cores)].add((l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes))

    n_added = 0
    n_expected = 0
    for ac, traces in per_ac_traces.items():
        configs = per_ac_configs[ac]
        for trace in traces:
            for cfg in configs:
                full_key = (ac[0], ac[1], trace) + cfg
                n_expected += 1
                if full_key not in data:
                    data[full_key] = {"ipc": SENTINEL_RUNNING}
                    n_added += 1
    return n_added, n_expected


def validate_ipc_count(outputs_dir, python_ok_count):
    """Cross-check python's IPC extraction against independent grep.

    Returns (grep_count, missing_basenames). Warns if python undercounted.
    """
    try:
        # grep -c per file, then sum; also capture list of files that have the marker
        result = subprocess.run(
            [
                "bash",
                "-c",
                f"grep -rl 'FINAL ROI CORE AVG IPC' {outputs_dir}/core*/ 2>/dev/null || true",
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
        files_with_ipc = [
            ln.strip() for ln in result.stdout.splitlines() if ln.strip()
        ]
    except (subprocess.SubprocessError, OSError):
        return None, []
    grep_count = len(files_with_ipc)
    return grep_count, files_with_ipc


NO_MODEL_TOKEN = "no"


def _is_no_model(l1m, l2m, llcm):
    return l1m == NO_MODEL_TOKEN and l2m == NO_MODEL_TOKEN and llcm == NO_MODEL_TOKEN


def _model_name(l1m, l2m, llcm):
    """Collapse to single name when L1==L2==LLC, else 'L1+L2+LLC'."""
    if l1m == l2m == llcm:
        return l1m
    return f"{l1m}+{l2m}+{llcm}"


def _model_hermes(l1m, l2m, llcm, hermes):
    """Model combo with HERMES/OCP suffix folded in."""
    base = _model_name(l1m, l2m, llcm)
    return base if hermes == "off" else f"{base}|{hermes}"


def build_sheets_triple(data, field_name):
    """Like build_sheets, but each cell holds a 3-tuple (L1D, L2C, LLC).
    Groups by (arch, cores, pf_combo).
    """
    combined = defaultdict(lambda: defaultdict(dict))
    combined_traces = set()
    combined_models = defaultdict(set)
    combined_pfs = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes), metrics in data.items():
        model_combo = _model_hermes(l1m, l2m, llcm, hermes)
        pf_combo = f"L1:{l1pf} L2:{l2pf} L3:{llcpf}"
        triple = tuple(
            metrics.get((lvl, field_name), SENTINEL_RUNNING) for lvl in LEVELS_TRIPLE
        )
        combined[(arch, cores, pf_combo)][model_combo][trace] = triple
        combined_traces.add(trace)
        combined_models[(arch, cores, pf_combo)].add(model_combo)
        combined_pfs[(arch, cores)].add(pf_combo)

    return {
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_models": combined_models,
        "combined_pfs": combined_pfs,
    }


def build_sheets_triple_covbyp(data):
    """Triple-cell builder for 'coverage with bypass crediting'.

    Value = max(load_miss - byp_issued, 0). The downstream coverage formula
    then computes (base - this)/base per-level, giving credit to bypass loads
    as covered misses. If either underlying value missing → SENTINEL_RUNNING.
    """
    per_model = defaultdict(lambda: defaultdict(dict))
    per_model_traces = defaultdict(set)
    combined = defaultdict(lambda: defaultdict(dict))
    combined_traces = set()
    combined_models = defaultdict(set)
    combined_pfs = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes), metrics in data.items():
        model_combo = _model_hermes(l1m, l2m, llcm, hermes)
        pf_combo = f"L1:{l1pf} L2:{l2pf} L3:{llcpf}"
        triple = []
        for lvl in LEVELS_TRIPLE:
            mv = metrics.get((lvl, "load_miss"))
            bv = metrics.get((lvl, "byp_issued"), 0)
            if mv is None:
                triple.append(SENTINEL_RUNNING)
            else:
                triple.append(max(mv - (bv or 0), 0))
        triple = tuple(triple)
        combined[(arch, cores, pf_combo)][model_combo][trace] = triple
        combined_traces.add(trace)
        combined_models[(arch, cores, pf_combo)].add(model_combo)
        combined_pfs[(arch, cores)].add(pf_combo)

    return {
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_models": combined_models,
        "combined_pfs": combined_pfs,
    }


def build_sheets(data, metric_key):
    """Build sheet data. Groups by (arch, cores, pf_combo).

    metric_key is either 'ipc' or a (lvl, field) tuple.
    """
    combined = defaultdict(lambda: defaultdict(dict))
    combined_traces = set()
    combined_models = defaultdict(set)
    combined_pfs = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes), metrics in data.items():
        model_combo = _model_hermes(l1m, l2m, llcm, hermes)
        pf_combo = f"L1:{l1pf} L2:{l2pf} L3:{llcpf}"
        val = metrics.get(metric_key, SENTINEL_RUNNING)
        combined[(arch, cores, pf_combo)][model_combo][trace] = val
        combined_traces.add(trace)
        combined_models[(arch, cores, pf_combo)].add(model_combo)
        combined_pfs[(arch, cores)].add(pf_combo)

    return {
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_models": combined_models,
        "combined_pfs": combined_pfs,
    }


def base_sort_key(pf_combo):
    """Sort key so that the 'base' pf combo lands LAST in each group.

    CP1 base = all pf = 'no' (L1:no L2:no LLC:no). Fallback: lexicographic
    last by prefetcher string (highest ASCII sorts last in plain sort).
    """
    is_all_no = pf_combo == "L1:no L2:no LLC:no"
    # Tuple: (is_all_no, pf_combo). is_all_no True -> 1 which is > 0, so
    # the all-no row naturally sorts after any False (0) entry under default
    # ascending sort. Tiebreak on name.
    return (1 if is_all_no else 0, pf_combo)


# ═══════════════════════════════════════════════════════════════════════════
# TRANSPOSED LAYOUT — IPC / HOSC / Coverage
# ═══════════════════════════════════════════════════════════════════════════

NO_PF_COMBO = "L1:no L2:no L3:no"
_TRANSPOSED_GAP_COLS = 5

# PF row order: no-pf FIRST, then specific order
_PF_ORDER = {
    NO_PF_COMBO: 0,
    "L1:no L2:spp L3:no": 1,
    "L1:no L2:Zion1 L3:no": 2,
    "L1:no L2:bingo_dpc3 L3:no": 3,
    "L1:no L2:berti_c L3:no": 4,
}


def _pf_sort_key_transposed(pf_combo):
    """no-pf FIRST, then spp, Zion1, bingo, berti, then alpha."""
    return (_PF_ORDER.get(pf_combo, 100), pf_combo)


def _discover_sub_columns(data):
    """Discover unique (model_name, hermes) combos → ordered sub-column list.

    Order: Base (no,off) → OCP modes (no,hermes), (no,hmp), (no,ttp) → Forward models (X,off).
    Returns list of (model_name, hermes, display_label) tuples.
    """
    combos = set()
    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes) in data.keys():
        mn = _model_name(l1m, l2m, llcm)
        combos.add((mn, hermes))

    base_combos = []     # (no, off)
    ocp_combos = []      # (no, hermes/hmp/ttp)
    fwd_combos = []      # (model, off)
    fwd_ocp_combos = []  # (model, hermes/hmp/ttp) — future-proof

    ocp_order = {"hermes": 0, "hmp": 1, "ttp": 2}

    for mn, h in sorted(combos):
        if mn == NO_MODEL_TOKEN and h == "off":
            base_combos.append((mn, h, "Base"))
        elif mn == NO_MODEL_TOKEN and h != "off":
            ocp_combos.append((mn, h, h.upper()))
        elif mn != NO_MODEL_TOKEN and h == "off":
            fwd_combos.append((mn, h, mn))
        else:
            fwd_ocp_combos.append((mn, h, f"{mn}|{h}"))

    ocp_combos.sort(key=lambda x: ocp_order.get(x[1], 99))
    fwd_combos.sort(key=lambda x: x[0])
    fwd_ocp_combos.sort(key=lambda x: (x[0], ocp_order.get(x[1], 99)))

    return base_combos + ocp_combos + fwd_combos + fwd_ocp_combos


def build_sheets_transposed(data, metric_key="ipc"):
    """Build transposed sheet data.

    Returns dict with:
      combined[(arch, cores, repl, pf_combo)][(model_name, hermes)][trace] = value
      traces: sorted list
      sub_cols: ordered list from _discover_sub_columns
      repls_per_ac: {(arch,cores): sorted list of repls}
      pfs_per_acr: {(arch,cores,repl): sorted list of pf_combos}
    """
    combined = defaultdict(lambda: defaultdict(dict))
    all_traces = set()
    repls_per_ac = defaultdict(set)
    pfs_per_acr = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes), metrics in data.items():
        mn = _model_name(l1m, l2m, llcm)
        pf_combo = f"L1:{l1pf} L2:{l2pf} L3:{llcpf}"
        val = metrics.get(metric_key, SENTINEL_RUNNING) if isinstance(metric_key, str) else metrics.get(metric_key, SENTINEL_RUNNING)
        combined[(arch, cores, repl, pf_combo)][(mn, hermes)][trace] = val
        all_traces.add(trace)
        repls_per_ac[(arch, cores)].add(repl)
        pfs_per_acr[(arch, cores, repl)].add(pf_combo)

    sub_cols = _discover_sub_columns(data)
    traces = sorted(all_traces)
    repls_sorted = {k: sorted(v) for k, v in repls_per_ac.items()}
    pfs_sorted = {k: sorted(v, key=_pf_sort_key_transposed) for k, v in pfs_per_acr.items()}

    return {
        "combined": combined,
        "traces": traces,
        "sub_cols": sub_cols,
        "repls_per_ac": repls_sorted,
        "pfs_per_acr": pfs_sorted,
    }


def _iter_transposed_groups(sh):
    """Yield ((arch, cores, repl), ordered_pf_combos) for transposed layout.

    Within each core count: one group per replacement policy.
    PF combos: no-pf FIRST, then ordered.
    Skip groups missing no-pf base.
    """
    ac_pairs = sorted({(a, c) for (a, c, r, p) in sh["combined"].keys()})
    for arch, cores in ac_pairs:
        repls = sh["repls_per_ac"].get((arch, cores), [])
        for repl in repls:
            pfs = sh["pfs_per_acr"].get((arch, cores, repl), [])
            if NO_PF_COMBO not in pfs:
                continue
            yield (arch, cores, repl), pfs


HEADER_COLS_TRANSPOSED = ["Cores", "Repl", "PF"]


def _write_transposed_header(ws, row, traces, sub_cols, first_trace_col):
    """Write 2-row header: row1 = merged trace names, row2 = sub-col labels.

    Also writes 5 gap cols between trace groups.
    Returns (next_row, trace_start_cols) where trace_start_cols[i] = first data col for trace i.
    """
    n_sub = len(sub_cols)
    trace_start_cols = []
    col = first_trace_col

    # Row 1: merged trace names
    for ti, trace in enumerate(traces):
        trace_start_cols.append(col)
        if n_sub > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_sub - 1)
        ws.cell(row=row, column=col, value=trace).alignment = Alignment(horizontal="center")
        col += n_sub
        col += _TRANSPOSED_GAP_COLS  # gap cols (empty)

    # GM and AVG merged headers
    gm_start = col
    if n_sub > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_sub - 1)
    ws.cell(row=row, column=col, value="GM").alignment = Alignment(horizontal="center")
    col += n_sub

    avg_start = col
    if n_sub > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_sub - 1)
    ws.cell(row=row, column=col, value="AVG").alignment = Alignment(horizontal="center")
    col += n_sub

    # Row 2: sub-col labels under each trace + GM + AVG
    row2 = row + 1
    col = first_trace_col
    for ti in range(len(traces)):
        for si, (mn, h, label) in enumerate(sub_cols):
            ws.cell(row=row2, column=col + si, value=label)
        col += n_sub + _TRANSPOSED_GAP_COLS

    # Sub-col labels under GM
    for si, (mn, h, label) in enumerate(sub_cols):
        ws.cell(row=row2, column=gm_start + si, value=label)
    # Sub-col labels under AVG
    for si, (mn, h, label) in enumerate(sub_cols):
        ws.cell(row=row2, column=avg_start + si, value=label)

    return row + 2, trace_start_cols, gm_start, avg_start


def _write_transposed_avg_gm(ws, row, trace_start_cols, sub_col_idx, n_sub, n_traces,
                              gm_start, avg_start, header_row=False, fmt="0.0000"):
    """Write GM and AVG for a single sub-column across all traces.

    Uses non-contiguous cell references (one per trace, same sub-col offset).
    """
    if header_row:
        return
    cols_for_sub = [trace_start_cols[ti] + sub_col_idx for ti in range(n_traces)]

    # GM formula: EXP(SUM(IF(valid,LN(x),0))/SUM(IF(valid,1,0)))
    gm_ln_parts = []
    gm_cnt_parts = []
    avg_sum_parts = []
    avg_cnt_parts = []
    for c in cols_for_sub:
        cl = get_column_letter(c)
        ref = f"{cl}{row}"
        gm_ln_parts.append(f'IF(AND({ref}<>-1,ISNUMBER({ref}),{ref}>0),LN({ref}),0)')
        gm_cnt_parts.append(f'IF(AND({ref}<>-1,ISNUMBER({ref}),{ref}>0),1,0)')
        avg_sum_parts.append(f'IF(AND({ref}<>-1,ISNUMBER({ref}),{ref}>0),{ref},0)')
        avg_cnt_parts.append(f'IF(AND({ref}<>-1,ISNUMBER({ref}),{ref}>0),1,0)')

    gm_f = f'=EXP(({"+".join(gm_ln_parts)})/({"+".join(gm_cnt_parts)}))'
    avg_f = f'=({"+".join(avg_sum_parts)})/({"+".join(avg_cnt_parts)})'

    cell_gm = ws.cell(row=row, column=gm_start + sub_col_idx, value=gm_f)
    cell_gm.number_format = fmt
    cell_avg = ws.cell(row=row, column=avg_start + sub_col_idx, value=avg_f)
    cell_avg.number_format = fmt


PF_DISPLAY = {
    "L1:no L2:no L3:no": "no Pref",
    "L1:no L2:spp L3:no": "SPP",
    "L1:no L2:Zion1 L3:no": "Zion",
    "L1:no L2:bingo_dpc3 L3:no": "Bingo",
    "L1:no L2:berti_c L3:no": "Berti",
}
AVGPF_LABEL = "AVG PF"


def _write_speedup_pivot(ws, anchor_row, anchor_col, title, row_specs,
                         sub_labels, gm_col_start, n_sub, fmt="0.0000"):
    """Linked pivot table placed to the RIGHT of the speedup GM/AVG region.

    row_specs: ordered (cores_label, repl_label, pf_label, gm_src_row). Each
    technique cell = '=<gm_col_letter><gm_src_row>' (same-sheet link to the GM
    column). A blank spacer row is emitted when the (cores,repl) group changes.
    Returns the column index just past this table.
    """
    ws.cell(row=anchor_row, column=anchor_col + 3, value=title)
    hr = anchor_row + 1
    ws.cell(row=hr, column=anchor_col + 0, value="Cores")
    ws.cell(row=hr, column=anchor_col + 1, value="Repl")
    ws.cell(row=hr, column=anchor_col + 2, value="PF")
    for j, lbl in enumerate(sub_labels):
        ws.cell(row=hr, column=anchor_col + 3 + j, value=lbl)
    r = hr + 1
    prev_group = None
    for cores_label, repl_label, pf_label, gm_src_row in row_specs:
        group = (cores_label, repl_label)
        if prev_group is not None and group != prev_group:
            r += 1
        ws.cell(row=r, column=anchor_col + 0, value=cores_label)
        ws.cell(row=r, column=anchor_col + 1, value=repl_label)
        ws.cell(row=r, column=anchor_col + 2, value=pf_label)
        for j in range(n_sub):
            src = f"{get_column_letter(gm_col_start + j)}{gm_src_row}"
            c = ws.cell(row=r, column=anchor_col + 3 + j, value=f"={src}")
            c.number_format = fmt
        prev_group = group
        r += 1
    return anchor_col + 3 + n_sub


def _write_transposed_block(ws, sh, start_row, section_title, formula_kind="speedup"):
    """Transposed IPC-style block writer. Returns next_row.

    Layout:
      Section title
      2-row header (merged trace names + sub-col labels)
      Per (arch, cores): per repl: rows of PFs with sub-cols per trace
      Gap between repl groups
      Larger gap between core count sections
      Then speedup mirror below
    """
    fmt_fn = _speedup_formula if formula_kind == "speedup" else _coverage_formula
    mirror_word = "SPEEDUP" if formula_kind == "speedup" else "COVERAGE"

    traces = sh["traces"]
    sub_cols = sh["sub_cols"]
    n_traces = len(traces)
    n_sub = len(sub_cols)
    n_label_cols = len(HEADER_COLS_TRANSPOSED)
    first_trace_col = n_label_cols + 1

    row = start_row
    ws.cell(row=row, column=1, value=section_title)
    row += 1

    # Write header
    header_start = row
    row, trace_start_cols, gm_start, avg_start = _write_transposed_header(
        ws, row, traces, sub_cols, first_trace_col
    )
    # Write label col headers on row2 of header
    ws.cell(row=header_start + 1, column=1, value="Cores")
    ws.cell(row=header_start + 1, column=2, value="Repl")
    ws.cell(row=header_start + 1, column=3, value="PF")

    main_row_of = {}    # (arch, cores, repl, pf_combo, sub_col_idx) -> row
    base_row_of = {}    # (arch, cores, repl) -> row of no-pf row
    base_sub_idx = None # index of Base sub-column

    # Find base sub-col index
    for si, (mn, h, label) in enumerate(sub_cols):
        if mn == NO_MODEL_TOKEN and h == "off":
            base_sub_idx = si
            break

    # Data rows
    ac_pairs = sorted({(a, c) for (a, c, r, p) in sh["combined"].keys()})
    for ac_i, (arch, cores) in enumerate(ac_pairs):
        if ac_i > 0:
            row += 2  # larger gap between core count sections

        repls = sh["repls_per_ac"].get((arch, cores), [])
        for repl_i, repl in enumerate(repls):
            pfs = sh["pfs_per_acr"].get((arch, cores, repl), [])
            if NO_PF_COMBO not in pfs:
                continue

            for pf_combo in pfs:
                ws.cell(row=row, column=1, value=cores)
                ws.cell(row=row, column=2, value=repl)
                ws.cell(row=row, column=3, value=pf_combo)

                for ti, trace in enumerate(traces):
                    for si, (mn, h, label) in enumerate(sub_cols):
                        col = trace_start_cols[ti] + si
                        val = sh["combined"].get(
                            (arch, cores, repl, pf_combo), {}
                        ).get((mn, h), {}).get(trace, SENTINEL_RUNNING)
                        if val == SENTINEL_RUNNING or val == SENTINEL_ERRORED or val == 0:
                            ws.cell(row=row, column=col, value="")
                        else:
                            _set_ipc_cell(ws, row, col, val)

                # GM/AVG per sub-col
                for si in range(n_sub):
                    _write_transposed_avg_gm(
                        ws, row, trace_start_cols, si, n_sub, n_traces,
                        gm_start, avg_start, fmt="0.0000"
                    )

                main_row_of[(arch, cores, repl, pf_combo)] = row
                if pf_combo == NO_PF_COMBO:
                    base_row_of[(arch, cores, repl)] = row
                row += 1

            row += 1  # gap between repl groups

    # ── Speedup mirror ──
    row += 2
    ws.cell(row=row, column=1, value=f"{mirror_word} — {section_title}")
    row += 1

    # Re-write header for speedup section
    sp_header_start = row
    row, sp_trace_start_cols, sp_gm_start, sp_avg_start = _write_transposed_header(
        ws, row, traces, sub_cols, first_trace_col
    )
    ws.cell(row=sp_header_start + 1, column=1, value="Cores")
    ws.cell(row=sp_header_start + 1, column=2, value="Repl")
    ws.cell(row=sp_header_start + 1, column=3, value="PF")

    sp_row_of = {}   # (cores, repl, pf_combo|AVGPF_LABEL) -> speedup GM row
    for ac_i, (arch, cores) in enumerate(ac_pairs):
        if ac_i > 0:
            row += 2

        repls = sh["repls_per_ac"].get((arch, cores), [])
        for repl_i, repl in enumerate(repls):
            pfs = sh["pfs_per_acr"].get((arch, cores, repl), [])
            if NO_PF_COMBO not in pfs:
                continue

            base_r = base_row_of.get((arch, cores, repl))
            pf_rows = {}

            for pf_combo in pfs:
                ws.cell(row=row, column=1, value=cores)
                ws.cell(row=row, column=2, value=repl)
                ws.cell(row=row, column=3, value=pf_combo)

                src_row = main_row_of.get((arch, cores, repl, pf_combo))

                for ti in range(n_traces):
                    for si in range(n_sub):
                        col = trace_start_cols[ti] + si
                        cl = get_column_letter(col)
                        if base_r is None or src_row is None:
                            ws.cell(row=row, column=col, value="")
                        else:
                            # Speedup: src_subcol / base_row_Base_subcol
                            src_cl = cl
                            if base_sub_idx is not None:
                                base_col = trace_start_cols[ti] + base_sub_idx
                                base_cl = get_column_letter(base_col)
                            else:
                                base_cl = cl
                            formula = f'=IF(AND({src_cl}{src_row}>0,{base_cl}{base_r}>0),{src_cl}{src_row}/{base_cl}{base_r},"")'
                            cell = ws.cell(row=row, column=col, value=formula)
                            cell.number_format = "0.0000"

                # GM/AVG for speedup rows
                for si in range(n_sub):
                    _write_transposed_avg_gm(
                        ws, row, trace_start_cols, si, n_sub, n_traces,
                        sp_gm_start, sp_avg_start, fmt="0.0000"
                    )

                pf_rows[pf_combo] = row
                sp_row_of[(cores, repl, pf_combo)] = row
                row += 1

            # AVG PF end row: per-trace mean of the prefetcher speedup rows
            # (excludes no-Pref); its GM = geomean-over-traces of those means.
            pref_rows = [pf_rows[p] for p in pfs if p != NO_PF_COMBO and p in pf_rows]
            ws.cell(row=row, column=1, value=cores)
            ws.cell(row=row, column=2, value=repl)
            ws.cell(row=row, column=3, value=AVGPF_LABEL)
            if pref_rows:
                for ti in range(n_traces):
                    for si in range(n_sub):
                        col = trace_start_cols[ti] + si
                        cl = get_column_letter(col)
                        refs = ",".join(f"{cl}{pr}" for pr in pref_rows)
                        cell = ws.cell(row=row, column=col, value=f'=IFERROR(AVERAGE({refs}),"")')
                        cell.number_format = "0.0000"
                for si in range(n_sub):
                    _write_transposed_avg_gm(
                        ws, row, trace_start_cols, si, n_sub, n_traces,
                        sp_gm_start, sp_avg_start, fmt="0.0000"
                    )
            sp_row_of[(cores, repl, AVGPF_LABEL)] = row
            row += 1
            row += 1

    # ── Linked pivot tables to the RIGHT of the GM/AVG region ──
    sub_labels = [lbl for (_mn, _h, lbl) in sub_cols]

    def _compact_specs(want_repls):
        specs = []
        for (arch, c) in ac_pairs:
            clbl = f"{c} Core"
            for rp in sh["repls_per_ac"].get((arch, c), []):
                if rp not in want_repls:
                    continue
                br = sp_row_of.get((c, rp, NO_PF_COMBO))
                ar = sp_row_of.get((c, rp, AVGPF_LABEL))
                if br is None or ar is None:
                    continue
                specs.append((clbl, rp, "Base", br))
                specs.append((clbl, rp, "Avg", ar))
        return specs

    def _allpf_specs(want_repls):
        specs = []
        for (arch, c) in ac_pairs:
            clbl = f"{c} Core"
            for rp in sh["repls_per_ac"].get((arch, c), []):
                if rp not in want_repls:
                    continue
                for pf in sh["pfs_per_acr"].get((arch, c, rp), []):
                    rr = sp_row_of.get((c, rp, pf))
                    if rr is None:
                        continue
                    specs.append((clbl, rp, PF_DISPLAY.get(pf, pf), rr))
        return specs

    all_repls = []
    for (arch, c) in ac_pairs:
        for rp in sh["repls_per_ac"].get((arch, c), []):
            if rp not in all_repls:
                all_repls.append(rp)
    non_lru = [rp for rp in all_repls if rp != "lru"]
    lru_only = [rp for rp in all_repls if rp == "lru"]

    piv_col = sp_avg_start + n_sub + 4
    for title, specs in (
        ("AVG PF GM ALL", _compact_specs(all_repls)),
        ("Base/Avg — non-LRU", _compact_specs(non_lru)),
        ("Base/Avg — LRU", _compact_specs(lru_only)),
        ("All-PF — LRU", _allpf_specs(lru_only)),
    ):
        if not specs:
            continue
        piv_col = _write_speedup_pivot(
            ws, sp_header_start, piv_col, title, specs,
            sub_labels, sp_gm_start, n_sub) + 4

    return row


def _write_transposed_coverage_block(ws, sh, start_row, section_title, base_sub_idx=0):
    """Transposed coverage block: (base_subcol - model_subcol) / base_subcol.

    Coverage is computed WITHIN each row: each sub-col's value relative to
    the Base sub-col of the SAME row and trace. Used for HOSC and Load coverage.
    Returns next_row.
    """
    traces = sh["traces"]
    sub_cols = sh["sub_cols"]
    n_traces = len(traces)
    n_sub = len(sub_cols)
    n_label_cols = len(HEADER_COLS_TRANSPOSED)
    first_trace_col = n_label_cols + 1

    row = start_row

    # Raw data block first (same as transposed IPC)
    ws.cell(row=row, column=1, value=f"RAW — {section_title}")
    row += 1
    header_start = row
    row, trace_start_cols, gm_start, avg_start = _write_transposed_header(
        ws, row, traces, sub_cols, first_trace_col
    )
    ws.cell(row=header_start + 1, column=1, value="Cores")
    ws.cell(row=header_start + 1, column=2, value="Repl")
    ws.cell(row=header_start + 1, column=3, value="PF")

    ac_pairs = sorted({(a, c) for (a, c, r, p) in sh["combined"].keys()})
    for ac_i, (arch, cores) in enumerate(ac_pairs):
        if ac_i > 0:
            row += 2
        repls = sh["repls_per_ac"].get((arch, cores), [])
        for repl in repls:
            pfs = sh["pfs_per_acr"].get((arch, cores, repl), [])
            if NO_PF_COMBO not in pfs:
                continue
            for pf_combo in pfs:
                ws.cell(row=row, column=1, value=cores)
                ws.cell(row=row, column=2, value=repl)
                ws.cell(row=row, column=3, value=pf_combo)
                for ti, trace in enumerate(traces):
                    for si, (mn, h, label) in enumerate(sub_cols):
                        col = trace_start_cols[ti] + si
                        val = sh["combined"].get(
                            (arch, cores, repl, pf_combo), {}
                        ).get((mn, h), {}).get(trace, SENTINEL_RUNNING)
                        if val == SENTINEL_RUNNING or val == SENTINEL_ERRORED or val == 0:
                            ws.cell(row=row, column=col, value="")
                        else:
                            _set_ipc_cell(ws, row, col, val)
                for si in range(n_sub):
                    _write_transposed_avg_gm(
                        ws, row, trace_start_cols, si, n_sub, n_traces,
                        gm_start, avg_start, fmt="0.0000"
                    )
                row += 1
            row += 1

    # Coverage mirror: (base_subcol - this_subcol) / base_subcol per row
    row += 2
    ws.cell(row=row, column=1, value=f"COVERAGE — {section_title}")
    row += 1
    cov_header_start = row
    row, cov_trace_start_cols, cov_gm_start, cov_avg_start = _write_transposed_header(
        ws, row, traces, sub_cols, first_trace_col
    )
    ws.cell(row=cov_header_start + 1, column=1, value="Cores")
    ws.cell(row=cov_header_start + 1, column=2, value="Repl")
    ws.cell(row=cov_header_start + 1, column=3, value="PF")

    # Re-iterate to write coverage formulas referencing the raw block above
    raw_row_of = {}
    raw_row = header_start + 2  # first data row in raw block
    for ac_i, (arch, cores) in enumerate(ac_pairs):
        if ac_i > 0:
            raw_row += 2
        repls = sh["repls_per_ac"].get((arch, cores), [])
        for repl in repls:
            pfs = sh["pfs_per_acr"].get((arch, cores, repl), [])
            if NO_PF_COMBO not in pfs:
                continue
            for pf_combo in pfs:
                raw_row_of[(arch, cores, repl, pf_combo)] = raw_row
                raw_row += 1
            raw_row += 1

    for ac_i, (arch, cores) in enumerate(ac_pairs):
        if ac_i > 0:
            row += 2
        repls = sh["repls_per_ac"].get((arch, cores), [])
        for repl in repls:
            pfs = sh["pfs_per_acr"].get((arch, cores, repl), [])
            if NO_PF_COMBO not in pfs:
                continue
            for pf_combo in pfs:
                ws.cell(row=row, column=1, value=cores)
                ws.cell(row=row, column=2, value=repl)
                ws.cell(row=row, column=3, value=pf_combo)
                src_raw_row = raw_row_of.get((arch, cores, repl, pf_combo))
                for ti in range(n_traces):
                    base_col = trace_start_cols[ti] + base_sub_idx
                    base_cl = get_column_letter(base_col)
                    for si in range(n_sub):
                        col = trace_start_cols[ti] + si
                        cl = get_column_letter(col)
                        if src_raw_row is None:
                            ws.cell(row=row, column=col, value="")
                        else:
                            # Coverage = (base - this) / base
                            formula = (
                                f'=IF(AND({base_cl}{src_raw_row}>0,{cl}{src_raw_row}>0,'
                                f'ISNUMBER({base_cl}{src_raw_row}),ISNUMBER({cl}{src_raw_row})),'
                                f'({base_cl}{src_raw_row}-{cl}{src_raw_row})/{base_cl}{src_raw_row},"")'
                            )
                            cell = ws.cell(row=row, column=col, value=formula)
                            cell.number_format = "0.00%"
                for si in range(n_sub):
                    _write_transposed_avg_gm(
                        ws, row, trace_start_cols, si, n_sub, n_traces,
                        cov_gm_start, cov_avg_start, fmt="0.00%"
                    )
                row += 1
            row += 1

    return row


HEADER_COLS_PER_MODEL = ["Arch", "Cores", "Pf \\ Traces"]
HEADER_COLS_COMBINED = ["Arch", "Cores", "Model", "PF \\ Traces"]


def _avg_formula(first_col_letter, last_col_letter, row):
    """AVERAGEIF excluding blanks, 0, and -1 sentinels."""
    rng = f"${first_col_letter}{row}:${last_col_letter}{row}"
    return f'=AVERAGEIF({rng},">0")'


def _gm_formula(first_col_letter, last_col_letter, row):
    """Geometric mean excluding -1 and <=0."""
    rng = f"${first_col_letter}{row}:${last_col_letter}{row}"
    return (
        f'=EXP(SUMPRODUCT(IF(({rng}<>-1)*ISNUMBER({rng})*({rng}>0),'
        f'LN({rng}),0))/SUMPRODUCT(({rng}<>-1)*ISNUMBER({rng})*({rng}>0)))'
    )


import math


def _topn_avg_indices(values, n):
    """Return indices of top-N traces for best AVG (remove worst speedups first).
    values: list of (index, float). Filters out sentinels (<0). Returns sorted indices."""
    valid = [(i, v) for i, v in values if isinstance(v, (int, float)) and v >= 0]
    if len(valid) <= n:
        return [i for i, _ in valid]
    # Sort by value ascending, remove worst (total - n)
    valid.sort(key=lambda x: x[1])
    keep = valid[len(valid) - n:]
    return sorted(i for i, _ in keep)


def _topn_gm_indices(values, n):
    """Return indices of top-N traces for best GM (greedy removal of worst GM contributor).
    values: list of (index, float). Filters out sentinels. Returns sorted indices."""
    valid = [(i, v) for i, v in values if isinstance(v, (int, float)) and v > 0]
    if len(valid) <= n:
        return [i for i, _ in valid]
    remaining = list(valid)
    while len(remaining) > n:
        # Find which removal maximizes GM of the rest
        best_remove = 0
        best_gm = -1
        log_sum = sum(math.log(v) for _, v in remaining)
        cnt = len(remaining)
        for ri in range(len(remaining)):
            # GM without this element
            new_log_sum = log_sum - math.log(remaining[ri][1])
            new_gm = math.exp(new_log_sum / (cnt - 1))
            if new_gm > best_gm:
                best_gm = new_gm
                best_remove = ri
        remaining.pop(best_remove)
    return sorted(i for i, _ in remaining)


def _compute_avg(values):
    valid = [v for v in values if isinstance(v, (int, float)) and v >= 0]
    return sum(valid) / len(valid) if valid else 0


def _compute_gm(values):
    valid = [v for v in values if isinstance(v, (int, float)) and v > 0]
    if not valid:
        return 0
    return math.exp(sum(math.log(v) for v in valid) / len(valid))


DEFAULT_TOPN = 40


def _write_avg_gm(ws, row, gap_col, first_col, last_col, header_row=False, fmt="0.0000"):
    """Write AVG and GM cells at gap_col+1 and gap_col+2.
    If header_row, write labels instead of formulas."""
    if header_row:
        ws.cell(row=row, column=gap_col + 1, value="AVG")
        ws.cell(row=row, column=gap_col + 2, value="GM")
        return
    fc = get_column_letter(first_col)
    lc = get_column_letter(last_col)
    cell_avg = ws.cell(row=row, column=gap_col + 1, value=_avg_formula(fc, lc, row))
    cell_avg.number_format = fmt
    cell_gm = ws.cell(row=row, column=gap_col + 2, value=_gm_formula(fc, lc, row))
    cell_gm.number_format = fmt


def _write_triple_avg_gm(ws, row, gap_col, first_trace_col, n_traces, header_row=False, fmt="0.0000"):
    """For triple sheets: write AVG/GM per level (L1D, L2C, LLC).
    Picks every 3rd cell starting at offset 0,1,2 respectively.
    Uses AVERAGEIF on the specific level columns only.
    Layout after gap_col: [L1D_AVG, L1D_GM, L2C_AVG, L2C_GM, LLC_AVG, LLC_GM]
    """
    if header_row:
        for li, lvl in enumerate(("L1D", "L2C", "LLC")):
            ws.cell(row=row, column=gap_col + 1 + li * 2, value=f"{lvl}_AVG")
            ws.cell(row=row, column=gap_col + 2 + li * 2, value=f"{lvl}_GM")
        return
    for li in range(3):
        # Build comma-separated cell refs for this level across all traces
        cols = [first_trace_col + ti * 3 + li for ti in range(n_traces)]
        # Use a helper range approach: reference individual cells
        # Since they're not contiguous, build with AVERAGE of IF array
        refs = ",".join(get_column_letter(c) + str(row) for c in cols)
        first_c = get_column_letter(cols[0])
        last_c = get_column_letter(cols[-1])
        # For non-contiguous cols, we use array formula scanning the full range
        # and picking every 3rd col via MOD. Simpler: just use the full triple
        # range and filter by column position.
        # Actually easiest: build explicit cell list
        avg_parts = []
        gm_ln_parts = []
        gm_cnt_parts = []
        for c in cols:
            cl = get_column_letter(c)
            ref = f"{cl}{row}"
            avg_parts.append(f'IF({ref}<>-1,{ref},FALSE)')
            gm_ln_parts.append(f'IF(AND({ref}<>-1,{ref}>0),LN({ref}),0)')
            gm_cnt_parts.append(f'IF(AND({ref}<>-1,{ref}>0),1,0)')
        # AVERAGE with IF: =AVERAGE(IF(c1<>-1,c1,FALSE),IF(c2<>-1,...))
        # openpyxl: plain formula. For many cols this gets long but works.
        # Simpler: AVERAGEIFS won't work on non-contiguous. Use helper approach:
        # just sum valid / count valid
        sum_parts = "+".join(f'IF({get_column_letter(c)}{row}<>-1,{get_column_letter(c)}{row},0)' for c in cols)
        cnt_parts = "+".join(f'IF({get_column_letter(c)}{row}<>-1,1,0)' for c in cols)
        avg_f = f'=({sum_parts})/({cnt_parts})'
        gm_ln = "+".join(gm_ln_parts)
        gm_cnt = "+".join(gm_cnt_parts)
        gm_f = f'=EXP(({gm_ln})/({gm_cnt}))'

        cell_avg = ws.cell(row=row, column=gap_col + 1 + li * 2, value=avg_f)
        cell_avg.number_format = fmt
        cell_gm = ws.cell(row=row, column=gap_col + 2 + li * 2, value=gm_f)
        cell_gm.number_format = fmt


def _col_letter(col_idx):
    s = ""
    n = col_idx
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def _write_header(ws, row, labels, traces):
    for i, lab in enumerate(labels):
        ws.cell(row=row, column=1 + i, value=lab)
    start = len(labels) + 1
    for i, trace in enumerate(traces):
        ws.cell(row=row, column=start + i, value=trace)


def _set_ipc_cell(ws, row, col, val):
    cell = ws.cell(row=row, column=col, value=val)
    if isinstance(val, float) and val >= 0:
        cell.number_format = "0.0000"
    return cell


def _speedup_formula(src_col_letter, src_row, base_row):
    return (
        f'=IF(AND({src_col_letter}{src_row}>=0,{src_col_letter}{base_row}>0),'
        f'{src_col_letter}{src_row}/{src_col_letter}{base_row},"")'
    )


def _coverage_formula(src_col_letter, src_row, base_row):
    # (base - src) / base — fraction of misses avoided vs no-model base
    return (
        f'=IF(AND({src_col_letter}{src_row}>=0,{src_col_letter}{base_row}>0),'
        f'({src_col_letter}{base_row}-{src_col_letter}{src_row})'
        f'/{src_col_letter}{base_row},"")'
    )


def _row_all_missing(sh, arch, cores, pf_combo, model_combo, traces):
    """True if every trace value for this row is SENTINEL_RUNNING (-1)."""
    for trace in traces:
        val = sh["combined"][(arch, cores)][pf_combo][model_combo].get(trace, SENTINEL_RUNNING)
        if isinstance(val, tuple):
            # triple-cell: check if any sub-value is not sentinel
            if any(v != SENTINEL_RUNNING for v in val):
                return False
        elif val != SENTINEL_RUNNING:
            return False
    return True


NO_PF_COMBO = "L1:no L2:no L3:no"


def _pf_group_sort_key(key):
    """Sort (arch, cores, pf_combo) so no-PF group comes FIRST per (arch,cores)."""
    arch, cores, pf_combo = key
    return (arch, cores, 0 if pf_combo == NO_PF_COMBO else 1, pf_combo)


def _iter_groups(sh):
    """Yield ((arch, cores, model_combo), ordered_pf_combos) for every data group.

    Groups by (arch, cores, model_combo). 'no' model first, then others.
    PF combos sorted with NO_PF_COMBO LAST (base row for speedup formula).
    Skip groups where no-PF base is missing.
    """
    # Pivot: collect all (arch, cores, model) combos and their PF sets
    from collections import defaultdict as _dd
    model_groups = _dd(set)
    for (arch, cores, pf_combo) in sh["combined"].keys():
        for model_combo in sh["combined"][(arch, cores, pf_combo)].keys():
            model_groups[(arch, cores, model_combo)].add(pf_combo)

    def _model_sort_key(key):
        arch, cores, model = key
        return (arch, cores, 0 if model == NO_MODEL_TOKEN else 1, model)

    for (arch, cores, model_combo) in sorted(model_groups.keys(), key=_model_sort_key):
        pf_combos = sorted(model_groups[(arch, cores, model_combo)])
        non_base = [p for p in pf_combos if p != NO_PF_COMBO]
        base_pfs = [p for p in pf_combos if p == NO_PF_COMBO]
        if not base_pfs:
            continue
        ordered = non_base + base_pfs
        yield (arch, cores, model_combo), ordered


def _write_single_block(
    ws,
    sh,
    start_row,
    formula_kind,
    section_title,
):
    """Single-cell-per-trace block (used by IPC). Returns next_row.

    Simplified: no PF, no LIM/rank. -1 → empty.
    Layout: title → groups by (arch,cores): header, data rows, gap.
    Then speedup mirror with AVG/GM.
    """
    fmt_fn = _speedup_formula if formula_kind == "speedup" else _coverage_formula
    mirror_word = "SPEEDUP" if formula_kind == "speedup" else "COVERAGE"
    traces = sorted(sh["combined_traces"])
    n_traces = len(traces)
    n_label_cols = len(HEADER_COLS_COMBINED)
    first_trace_col = n_label_cols + 1
    last_trace_col = first_trace_col + n_traces - 1
    gap_col = last_trace_col + 1

    row = start_row
    ws.cell(row=row, column=1, value=section_title)
    row += 1

    main_row_of = {}
    base_row_of = {}

    for (arch, cores, model_combo), ordered_pfs in _iter_groups(sh):
        _write_header(ws, row, HEADER_COLS_COMBINED, traces)
        _write_avg_gm(ws, row, gap_col, first_trace_col, last_trace_col, header_row=True)
        row += 1
        for pf_combo in ordered_pfs:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=model_combo)
            ws.cell(row=row, column=4, value=pf_combo)
            for i, trace in enumerate(traces):
                val = sh["combined"][(arch, cores, pf_combo)][model_combo].get(
                    trace, SENTINEL_RUNNING
                )
                if val == SENTINEL_RUNNING or val == SENTINEL_ERRORED or val == 0:
                    ws.cell(row=row, column=first_trace_col + i, value="")
                else:
                    _set_ipc_cell(ws, row, first_trace_col + i, val)
            _write_avg_gm(ws, row, gap_col, first_trace_col, last_trace_col)
            main_row_of[(arch, cores, model_combo, pf_combo)] = row
            if pf_combo == NO_PF_COMBO:
                base_row_of[(arch, cores, model_combo)] = row
            row += 1
        row += 1

    # Speedup mirror
    row += 2
    ws.cell(row=row, column=1, value=f"{mirror_word} — {section_title}")
    row += 1

    for (arch, cores, model_combo), ordered_pfs in _iter_groups(sh):
        _write_header(ws, row, HEADER_COLS_COMBINED, traces)
        _write_avg_gm(ws, row, gap_col, first_trace_col, last_trace_col, header_row=True)
        row += 1
        base_r = base_row_of.get((arch, cores, model_combo))
        for pf_combo in ordered_pfs:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=model_combo)
            ws.cell(row=row, column=4, value=pf_combo)
            src_row = main_row_of[(arch, cores, model_combo, pf_combo)]
            for i in range(n_traces):
                col_idx = first_trace_col + i
                cl = get_column_letter(col_idx)
                if base_r is None:
                    ws.cell(row=row, column=col_idx, value="")
                else:
                    formula = fmt_fn(cl, src_row, base_r)
                    cell = ws.cell(row=row, column=col_idx, value=formula)
                    cell.number_format = "0.0000"
            _write_avg_gm(ws, row, gap_col, first_trace_col, last_trace_col)
            row += 1
        row += 1

    return row


def _write_metric_sheets(wb, sh, metric_label, formula_kind="speedup"):
    """Emit the single combined IPC-style sheet (no per-model sheets)."""
    combined_title = f"{metric_label} ALL (comb)"[:31]
    ws = wb.create_sheet(title=combined_title)
    _write_single_block(
        ws,
        sh,
        start_row=1,
        formula_kind=formula_kind,
        section_title=f"{metric_label}: all models combined (base = no-model per PF)",
    )


def _write_merged_single_sheet(wb, data_by_repl, repls, metric_label, metric_key,
                                formula_kind="speedup"):
    """Merged legacy IPC-style sheet: stack repl groups vertically in one sheet."""
    ws = wb.create_sheet(title=metric_label[:31])
    row = 1
    for repl in repls:
        rd = data_by_repl.get(repl, {})
        if not rd:
            continue
        sh = build_sheets(rd, metric_key)
        row = _write_single_block(
            ws, sh, start_row=row, formula_kind=formula_kind,
            section_title=f"{metric_label} — {repl}",
        )
        row += 2


def _write_merged_triple_sheet(wb, data_by_repl, repls, metric_label, field_name,
                                formula_kind="speedup"):
    """Merged legacy triple sheet: stack repl groups vertically in one sheet."""
    ws = wb.create_sheet(title=metric_label[:31])
    row = 1
    for repl in repls:
        rd = data_by_repl.get(repl, {})
        if not rd:
            continue
        sh = build_sheets_triple(rd, field_name)
        row = _write_triple_block(
            ws, sh, start_row=row, formula_kind=formula_kind,
            section_title=f"{metric_label} — {repl}",
        )
        row += 2


def _write_merged_coverage_sheet(wb, data_by_repl, repls, metric_label):
    """Merged legacy coverage sheet: stack repl groups vertically in one sheet."""
    ws_title = metric_label[:31]
    # Coverage uses _write_consolidated_coverage_sheet which creates its own sheets,
    # so we need a different approach: create temp wb per repl, merge
    from openpyxl import Workbook as _WB
    row_offset = 0
    ws = wb.create_sheet(title=ws_title)
    for repl in repls:
        rd = data_by_repl.get(repl, {})
        if not rd:
            continue
        tmp_wb = _WB()
        tmp_wb.remove(tmp_wb.active)
        _write_consolidated_coverage_sheet(tmp_wb, rd, title_override=f"Cov_{repl}")
        for tmp_ws in tmp_wb.worksheets:
            # Title row for repl
            ws.cell(row=row_offset + 1, column=1, value=f"Coverage — {repl}")
            row_offset += 1
            for src_row in tmp_ws.iter_rows():
                for cell in src_row:
                    dst_cell = ws.cell(
                        row=cell.row + row_offset,
                        column=cell.column,
                        value=cell.value,
                    )
                    if cell.has_style:
                        dst_cell.number_format = cell.number_format
            for mcr in tmp_ws.merged_cells.ranges:
                # Offset merged range rows
                from openpyxl.utils import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(str(mcr))
                ws.merge_cells(
                    start_row=min_row + row_offset,
                    start_column=min_col,
                    end_row=max_row + row_offset,
                    end_column=max_col,
                )
            row_offset += tmp_ws.max_row + 2


def _write_triple_group_header(ws, row, traces, first_trace_col):
    """Two-row header: row N has label cols + merged trace names,
    row N+1 has L1D/L2C/LLC sub-labels under each trace."""
    for i, lab in enumerate(HEADER_COLS_COMBINED):
        ws.cell(row=row, column=1 + i, value=lab)
    for ti, trace in enumerate(traces):
        c0 = first_trace_col + ti * 3
        cell = ws.cell(row=row, column=c0, value=trace)
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 2)
    for ti in range(len(traces)):
        c0 = first_trace_col + ti * 3
        for li, lvl in enumerate(LEVELS_TRIPLE):
            ws.cell(row=row + 1, column=c0 + li, value=lvl)


def _write_triple_block(ws, sh, start_row, formula_kind, section_title):
    """Triple-cell (L1D/L2C/LLC) block. Simplified: no PF, no LIM. -1→empty."""
    fmt_fn = _speedup_formula if formula_kind == "speedup" else _coverage_formula
    mirror_word = "SPEEDUP" if formula_kind == "speedup" else "COVERAGE"
    traces = sorted(sh["combined_traces"])
    n_traces = len(traces)
    n_label_cols = len(HEADER_COLS_COMBINED)
    first_trace_col = n_label_cols + 1
    last_triple_col = first_trace_col + n_traces * 3 - 1
    gap_col = last_triple_col + 1

    row = start_row
    ws.cell(row=row, column=1, value=section_title)
    row += 1

    main_row_of = {}
    base_row_of = {}

    for (arch, cores, model_combo), ordered_pfs in _iter_groups(sh):
        _write_triple_group_header(ws, row, traces, first_trace_col)
        _write_triple_avg_gm(ws, row + 1, gap_col, first_trace_col, n_traces, header_row=True)
        row += 2
        for pf_combo in ordered_pfs:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=model_combo)
            ws.cell(row=row, column=4, value=pf_combo)
            for ti, trace in enumerate(traces):
                triple = sh["combined"][(arch, cores, pf_combo)][model_combo].get(
                    trace, (SENTINEL_RUNNING,) * 3
                )
                c0 = first_trace_col + ti * 3
                for li in range(3):
                    v = triple[li]
                    if v == SENTINEL_RUNNING or v == SENTINEL_ERRORED or v == 0:
                        ws.cell(row=row, column=c0 + li, value="")
                    else:
                        cell = ws.cell(row=row, column=c0 + li, value=v)
                        if isinstance(v, float) and v >= 0:
                            cell.number_format = "0.0000"
            _write_triple_avg_gm(ws, row, gap_col, first_trace_col, n_traces)
            main_row_of[(arch, cores, model_combo, pf_combo)] = row
            if pf_combo == NO_PF_COMBO:
                base_row_of[(arch, cores, model_combo)] = row
            row += 1
        row += 1

    # Speedup mirror
    row += 2
    ws.cell(row=row, column=1, value=f"{mirror_word} — {section_title}")
    row += 1

    for (arch, cores, model_combo), ordered_pfs in _iter_groups(sh):
        _write_triple_group_header(ws, row, traces, first_trace_col)
        _write_triple_avg_gm(ws, row + 1, gap_col, first_trace_col, n_traces, header_row=True)
        row += 2
        base_r = base_row_of.get((arch, cores, model_combo))
        for pf_combo in ordered_pfs:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=model_combo)
            ws.cell(row=row, column=4, value=pf_combo)
            src_row = main_row_of[(arch, cores, model_combo, pf_combo)]
            for ti in range(n_traces):
                c0 = first_trace_col + ti * 3
                for li in range(3):
                    col = get_column_letter(c0 + li)
                    if base_r is None:
                        ws.cell(row=row, column=c0 + li, value="")
                    else:
                        formula = fmt_fn(col, src_row, base_r)
                        cell = ws.cell(row=row, column=c0 + li, value=formula)
                        cell.number_format = "0.0000"
            _write_triple_avg_gm(ws, row, gap_col, first_trace_col, n_traces)
            row += 1
        row += 1

    return row


def _write_triple_family_sheet(wb, sh, metric_label, formula_kind="speedup"):
    title = f"{metric_label} ALL (comb)"[:31]
    ws = wb.create_sheet(title=title)
    _write_triple_block(
        ws, sh, start_row=1, formula_kind=formula_kind,
        section_title=f"{metric_label}: per model, L1D/L2C/LLC triple",
    )


def build_sheets_multilevel(data, field_per_sub):
    """Multilevel builder — groups by (arch, cores, pf_combo)."""
    combined = defaultdict(lambda: defaultdict(dict))
    combined_traces = set()
    combined_models = defaultdict(set)
    combined_pfs = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm, repl, hermes), metrics in data.items():
        model_combo = _model_hermes(l1m, l2m, llcm, hermes)
        pf_combo = f"L1:{l1pf} L2:{l2pf} L3:{llcpf}"
        packed = []
        for lvl in LEVELS_TRIPLE:
            for f in field_per_sub:
                packed.append(metrics.get((lvl, f), SENTINEL_RUNNING))
        packed = tuple(packed)
        combined[(arch, cores, pf_combo)][model_combo][trace] = packed
        combined_traces.add(trace)
        combined_models[(arch, cores, pf_combo)].add(model_combo)
        combined_pfs[(arch, cores)].add(pf_combo)

    return {
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_models": combined_models,
        "combined_pfs": combined_pfs,
    }


def _write_multilevel_header(ws, row, traces, first_trace_col, sub_labels):
    """Three-row header: merged trace names, merged level names under each
    trace (spanning len(sub_labels) cells each), and sub-labels under the
    levels.
    """
    width = 3 * len(sub_labels)
    for i, lab in enumerate(HEADER_COLS_COMBINED):
        ws.cell(row=row, column=1 + i, value=lab)
    for ti, trace in enumerate(traces):
        c0 = first_trace_col + ti * width
        cell = ws.cell(row=row, column=c0, value=trace)
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=row, start_column=c0, end_row=row, end_column=c0 + width - 1
        )
    # row+1: level labels (L1D / L2C / LLC), each spanning len(sub_labels) cols
    for ti in range(len(traces)):
        c0 = first_trace_col + ti * width
        for li, lvl in enumerate(LEVELS_TRIPLE):
            sub_c0 = c0 + li * len(sub_labels)
            cell = ws.cell(row=row + 1, column=sub_c0, value=lvl)
            cell.alignment = Alignment(horizontal="center")
            if len(sub_labels) > 1:
                ws.merge_cells(
                    start_row=row + 1,
                    start_column=sub_c0,
                    end_row=row + 1,
                    end_column=sub_c0 + len(sub_labels) - 1,
                )
    # row+2: sub-labels repeated under each level
    for ti in range(len(traces)):
        c0 = first_trace_col + ti * width
        for li in range(3):
            for si, sub_lab in enumerate(sub_labels):
                ws.cell(
                    row=row + 2,
                    column=c0 + li * len(sub_labels) + si,
                    value=sub_lab,
                )


def _write_coverage_multicol_block(ws, data, start_row, field_per_sub, sub_labels,
                                   section_title, percent_format=True):
    """Emit a coverage block where each trace header spans 3*len(sub_labels) cols.

    Main region: raw counts. Mirror region: coverage formula (base-this)/base
    applied ONLY to sub-columns whose label equals 'Misses'; the rest are blank.
    Percent format on mirror region cells when percent_format=True.
    """
    sh = build_sheets_multilevel(data, field_per_sub)
    traces = sorted(sh["combined_traces"])
    n_sub = len(sub_labels)
    width_per_trace = 3 * n_sub
    n_label_cols = len(HEADER_COLS_COMBINED)
    first_trace_col = n_label_cols + 1

    row = start_row
    ws.cell(row=row, column=1, value=section_title)
    row += 1

    main_row_of = {}
    base_row_of = {}

    for (arch, cores, model_combo), ordered_pfs in _iter_groups(sh):
        _write_multilevel_header(ws, row, traces, first_trace_col, sub_labels)
        row += 3
        for pf_combo in ordered_pfs:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=model_combo)
            ws.cell(row=row, column=4, value=pf_combo)
            for ti, trace in enumerate(traces):
                packed = sh["combined"][(arch, cores, pf_combo)][model_combo].get(
                    trace, (SENTINEL_RUNNING,) * (3 * n_sub)
                )
                c0 = first_trace_col + ti * width_per_trace
                for idx in range(3 * n_sub):
                    v = packed[idx]
                    if v == SENTINEL_RUNNING or v == SENTINEL_ERRORED or v == 0:
                        ws.cell(row=row, column=c0 + idx, value="")
                    else:
                        cell = ws.cell(row=row, column=c0 + idx, value=v)
                        if isinstance(v, (int, float)) and v >= 0:
                            cell.number_format = "0"
            main_row_of[(arch, cores, model_combo, pf_combo)] = row
            if pf_combo == NO_PF_COMBO:
                base_row_of[(arch, cores, model_combo)] = row
            row += 1
        row += 1

    # Coverage mirror
    row += 2
    ws.cell(row=row, column=1, value=f"COVERAGE — {section_title}  (=(base-eval)/base)")
    row += 1

    miss_sub_idxs = [i for i, s in enumerate(sub_labels) if s.lower().startswith("miss")]
    byp_sub_idxs = [i for i, s in enumerate(sub_labels) if s.lower() == "byp"]
    miss_relative_idx = miss_sub_idxs[0] if miss_sub_idxs else None

    for (arch, cores, model_combo), ordered_pfs in _iter_groups(sh):
        _write_multilevel_header(ws, row, traces, first_trace_col, sub_labels)
        row += 3
        base_r = base_row_of.get((arch, cores, model_combo))
        for pf_combo in ordered_pfs:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=model_combo)
            ws.cell(row=row, column=4, value=pf_combo)
            src_row = main_row_of[(arch, cores, model_combo, pf_combo)]
            for ti in range(len(traces)):
                c0 = first_trace_col + ti * width_per_trace
                for li in range(3):
                    for si in range(n_sub):
                        idx = c0 + li * n_sub + si
                        col = get_column_letter(idx)
                        if si in miss_sub_idxs and base_r is not None:
                            formula = _coverage_formula(col, src_row, base_r)
                            cell = ws.cell(row=row, column=idx, value=formula)
                            cell.number_format = "0.00%" if percent_format else "0.0000"
                        elif si in byp_sub_idxs and base_r is not None and miss_relative_idx is not None:
                            byp_col = get_column_letter(idx)
                            base_miss_idx = c0 + li * n_sub + miss_relative_idx
                            base_miss_col = get_column_letter(base_miss_idx)
                            formula = (
                                f'=IF(AND({byp_col}{src_row}>=0,'
                                f'{base_miss_col}{base_r}>0),'
                                f'({base_miss_col}{base_r}-{byp_col}{src_row})'
                                f'/{base_miss_col}{base_r},"")'
                            )
                            cell = ws.cell(row=row, column=idx, value=formula)
                            cell.number_format = "0.00%" if percent_format else "0.0000"
                        else:
                            ws.cell(row=row, column=idx, value="")
            row += 1
        row += 1

    return row


def _write_consolidated_coverage_sheet(wb, data, title_override=None):
    """Single Coverage sheet with two sibling blocks:

    1) ByP-aware: per trace header → 9 cols = L1D/L2C/LLC × Hits/ByP/Misses,
       sourced from load_wByP_hit / load_wByP_byp / load_wByP_miss.
       Coverage formula ((base-eval)/base) on the Misses sub-cols.

    2) STD: per trace header → 6 cols = L1D/L2C/LLC × Hits/Misses,
       sourced from load_hit / load_miss. Coverage formula on Misses.

    Coverage displayed as percentage (0.00%).
    """
    ws = wb.create_sheet(title=(title_override or "Coverage")[:31])
    row = 1
    ws.cell(
        row=row,
        column=1,
        value="COVERAGE — ByP-aware (9-col: Hits/ByP/Misses per level), then STD (6-col: Hits/Misses per level)",
    )
    row += 2

    # Block 1: ByP-aware
    row = _write_coverage_multicol_block(
        ws,
        data,
        row,
        field_per_sub=["load_wByP_hit", "load_wByP_byp", "load_wByP_miss"],
        sub_labels=["Hits", "ByP", "Misses"],
        section_title="ByP-aware (load_wByP_hit / load_wByP_byp / load_wByP_miss)",
    )
    row += 3

    # Block 2: STD
    ws.cell(row=row, column=1, value="STANDARD (no ByP — reported load_hit / load_miss)")
    row += 2
    row = _write_coverage_multicol_block(
        ws,
        data,
        row,
        field_per_sub=["load_hit", "load_miss"],
        sub_labels=["Hits", "Misses"],
        section_title="STD (load_hit / load_miss)",
    )


_CF_HDR_LABELS = {
    "Model", "model", "Model \\ Traces", "[rank]", "Arch", "Cores", "PF",
    "MAX", "N_VALID", "CUTOFF", "LIM-AVG", "LIM-GM",
}


# Sheets where LOWER values are BETTER (red=high, green=low)
_LOWER_IS_BETTER_SHEETS = {
    "MPKI", "C-AMAT", "MST", "Miss", "LoadMSHRcap", "PfMSHRcap",
    "LPM_C-AMAT", "LPM_MST", "PureMiss", "Overlap",
}


def _apply_group_conditional_formatting(wb):
    """Apply color scale per data-group per column on final workbook.

    Higher-is-better sheets: red(low)→yellow(mid)→green(high).
    Lower-is-better sheets: green(low)→yellow(mid)→red(high).
    Groups detected by contiguous non-empty numeric data rows in col D (first trace col).
    """
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 3:
            continue

        # Determine color direction from sheet title
        lower_better = any(ws.title.startswith(lb) or ws.title == lb
                          for lb in _LOWER_IS_BETTER_SHEETS)
        if lower_better:
            color_lo, color_hi = "63BE7B", "F8696B"  # green=low, red=high
        else:
            color_lo, color_hi = "F8696B", "63BE7B"  # red=low, green=high

        # Detect groups: contiguous rows where col E (first trace col) has numeric data or formulas
        groups = []
        in_group = False
        grp_start = 0
        for r in range(1, max_row + 1):
            is_data = False
            for probe_col in range(5, min(max_col + 1, 10)):
                cell_v = ws.cell(row=r, column=probe_col).value
                if cell_v is None:
                    continue
                if isinstance(cell_v, (int, float)):
                    is_data = True
                    break
                s = str(cell_v).strip()
                if s.startswith("="):
                    is_data = True
                    break
                try:
                    float(s)
                    is_data = True
                    break
                except ValueError:
                    pass
            if is_data and not in_group:
                grp_start = r
                in_group = True
            elif not is_data and in_group:
                groups.append((grp_start, r - 1))
                in_group = False
        if in_group:
            groups.append((grp_start, max_row))

        for gs, ge in groups:
            if ge - gs < 1:
                continue
            for ci in range(4, max_col + 1):
                cl = get_column_letter(ci)
                rng = f"{cl}{gs}:{cl}{ge}"
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min", start_color=color_lo,
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color=color_hi,
                    ),
                )


def _sheet_job(args):
    """Worker: build one sheet, save to its own xlsx in sheets_dir. Returns (order, sheet_name, tmp_path)."""
    order, sheet_name, kind, data, sheets_dir = args
    wb = Workbook()
    wb.remove(wb.active)
    # sheet_name is the desired final tab title; pass it through as the
    # metric_label so inner writers don't generate a clashing title.
    if kind == "ipc":
        sh = build_sheets(data, "ipc")
        _write_metric_sheets(wb, sh, sheet_name)
    elif kind == "ipc_transposed":
        sh = build_sheets_transposed(data, "ipc")
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_transposed_block(ws, sh, start_row=1,
                                section_title="IPC (transposed: sub-cols per trace, rows per repl+PF)",
                                formula_kind="speedup")
    elif kind[0] == "hosc_transposed":
        _, level, field_name = kind
        sh = build_sheets_transposed(data, (level, field_name))
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_transposed_coverage_block(ws, sh, start_row=1,
                                         section_title=f"HOSC {level} (LOAD_MSHR_cap coverage)")
    elif kind[0] == "load_cov_transposed":
        _, level, field_name = kind
        sh = build_sheets_transposed(data, (level, field_name))
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_transposed_coverage_block(ws, sh, start_row=1,
                                         section_title=f"Load Coverage {level} (miss reduction)")
    elif kind[0] == "triple_transposed":
        _, level, field_name, sheet_label = kind
        sh = build_sheets_transposed(data, (level, field_name))
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_transposed_block(ws, sh, start_row=1,
                                section_title=f"{sheet_label} {level} (transposed: sub-cols per trace)",
                                formula_kind="speedup")
    elif kind[0] == "triple_merged":
        # L1D block on top, then L2C, then LLC — stacked, gap lines between.
        _, field_name, sheet_label = kind
        ws = wb.create_sheet(title=sheet_name[:31])
        row = 1
        for level in LEVELS_TRIPLE:
            sh = build_sheets_transposed(data, (level, field_name))
            row = _write_transposed_block(
                ws, sh, start_row=row,
                section_title=f"{sheet_label} {level} (transposed: sub-cols per trace)",
                formula_kind="speedup")
            row += 3  # gap lines before next level
    elif kind[0] == "hosc_merged":
        ws = wb.create_sheet(title=sheet_name[:31])
        row = 1
        for level in LEVELS_TRIPLE:
            sh = build_sheets_transposed(data, (level, "LOAD_MSHR_cap"))
            row = _write_transposed_coverage_block(
                ws, sh, start_row=row,
                section_title=f"HOSC {level} (LOAD_MSHR_cap coverage)")
            row += 3
    elif kind[0] == "loadcov_merged":
        ws = wb.create_sheet(title=sheet_name[:31])
        row = 1
        for level in LEVELS_TRIPLE:
            sh = build_sheets_transposed(data, (level, "load_miss"))
            row = _write_transposed_coverage_block(
                ws, sh, start_row=row,
                section_title=f"Load Coverage {level} (miss reduction)")
            row += 3
    elif kind[0] == "merged_single":
        _, metric_key, formula_kind_inner, data_by_repl, repls = kind
        _write_merged_single_sheet(wb, data_by_repl, repls, sheet_name, metric_key,
                                    formula_kind=formula_kind_inner)
    elif kind[0] == "merged_triple":
        _, field_name, formula_kind_inner, data_by_repl, repls = kind
        _write_merged_triple_sheet(wb, data_by_repl, repls, sheet_name, field_name,
                                    formula_kind=formula_kind_inner)
    elif kind[0] == "merged_coverage":
        _, data_by_repl, repls = kind
        _write_merged_coverage_sheet(wb, data_by_repl, repls, sheet_name)
    elif kind == "coverage":
        _write_consolidated_coverage_sheet(wb, data, title_override=sheet_name)
    elif kind[0] == "triple":
        _, field, _label, formula = kind
        sh = build_sheets_triple(data, field)
        _write_triple_family_sheet(wb, sh, sheet_name, formula_kind=formula)
    # Rename any tab whose title exceeds 31 chars or differs (safety)
    for ws in wb.worksheets:
        ws.title = sheet_name[:31]
    tmp_path = os.path.join(sheets_dir, f"{order:02d}_{sheet_name}.xlsx".replace("/", "_"))
    wb.save(tmp_path)
    return order, sheet_name, tmp_path


def _filter_data_by_repl(data, repl):
    """Return subset of data dict where key[9] == repl."""
    return {k: v for k, v in data.items() if k[9] == repl}


def write_xlsx(data, input_basename, n_workers=8):
    ts = int(time.time())
    sheets_dir = os.path.join(OUTPUT_DIR, f"sheets_{input_basename}_{ts}")
    os.makedirs(sheets_dir, exist_ok=True)

    # Distinct REPL values, deterministic order
    repls = sorted({k[9] for k in data.keys()})
    if not repls:
        repls = ["lru"]

    def _sname(base, repl):
        # Sheet names capped at 31 chars by Excel
        return f"{base}_{repl}"[:31]

    # Per-repl data slices (computed once, reused)
    data_by_repl = {r: _filter_data_by_repl(data, r) for r in repls}

    # Sheet order: grouped by metric, then per-repl. Each metric block does
    # IPC_<repl1> IPC_<repl2> ... then MPKI_<repl1> MPKI_<repl2> ... etc.
    jobs = []
    order = 0

    # ═══════════════════════════════════════════════════════════════════
    # NEW FORMAT (TRANSPOSED): all repls merged as row groups under each core count
    # ═══════════════════════════════════════════════════════════════════

    # IPC (single value per cell)
    jobs.append((order, "IPC", "ipc_transposed", data, sheets_dir))
    order += 1

    # HOSC: 3 levels (L1D top, then L2C, then LLC) stacked in ONE sheet
    jobs.append((order, "HOSC", ("hosc_merged",), data, sheets_dir))
    order += 1

    # Load Coverage: 3 levels stacked in ONE sheet
    jobs.append((order, "LoadCov", ("loadcov_merged",), data, sheets_dir))
    order += 1

    # Triple metrics (MPKI, Miss, etc.) — 3 levels stacked in ONE sheet each
    triple_specs = []
    triple_specs.append(("MPKI", "MPKI"))
    _CUT_FAMILIES = {"MPKI", "Loads", "PfMSHRcap", "BypReq", "BypIss"}
    for sheet_label, field in CP5_FAMILIES:
        if sheet_label in _CUT_FAMILIES:
            continue
        triple_specs.append((sheet_label, field))

    for sheet_label, field_name in triple_specs:
        jobs.append((order, sheet_label[:31],
                     ("triple_merged", field_name, sheet_label),
                     data, sheets_dir))
        order += 1

    # ═══════════════════════════════════════════════════════════════════
    # LEGACY FORMAT MERGED: all repls stacked vertically in one sheet
    # Disabled by default — slow and superseded by transposed sheets.
    # ═══════════════════════════════════════════════════════════════════
    BUILD_LEGACY = False
    if BUILD_LEGACY:
        # Legacy merged IPC
        jobs.append((order, "IPC_legacy",
                     ("merged_single", "ipc", "speedup", data_by_repl, repls),
                     {}, sheets_dir))
        order += 1

        # Legacy merged Coverage
        jobs.append((order, "Coverage_legacy",
                     ("merged_coverage", data_by_repl, repls),
                     {}, sheets_dir))
        order += 1

        # Legacy merged triple sheets (MPKI, LPM_*, Loads, etc.)
        for sheet_label, field_name in triple_specs:
            jobs.append((order, f"{sheet_label}_legacy"[:31],
                         ("merged_triple", field_name, "speedup", data_by_repl, repls),
                         {}, sheets_dir))
            order += 1

    # Run in parallel
    results = []
    with ProcessPoolExecutor(max_workers=n_workers) as executor:
        for res in executor.map(_sheet_job, jobs):
            print(f"  sheet done: {res[1]}", flush=True)
            results.append(res)

    print(f"sheets_dir={sheets_dir}", flush=True)

    # Merge all individual sheets into one final xlsx
    from openpyxl import load_workbook as _load
    from copy import copy as _copy
    results.sort(key=lambda x: x[0])
    final_wb = Workbook()
    final_wb.remove(final_wb.active)
    for _, sheet_name, tmp_path in results:
        tmp_wb = _load(tmp_path)
        for src_ws in tmp_wb.worksheets:
            dst_ws = final_wb.create_sheet(title=src_ws.title)
            for row_cells in src_ws.iter_rows():
                for cell in row_cells:
                    dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        dst_cell.number_format = cell.number_format
            # Copy merged cell ranges
            for mcr in src_ws.merged_cells.ranges:
                dst_ws.merge_cells(str(mcr))
        print(f"  merged: {sheet_name}", flush=True)

    _apply_group_conditional_formatting(final_wb)
    print("  conditional formatting applied", flush=True)

    import chart_figs
    fig_sheets = chart_figs.add_figure_sheets(final_wb, data)
    print(f"  figure sheets added: {fig_sheets}", flush=True)

    out_name = f"{input_basename}_CP7b_{ts}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_name)
    final_wb.save(out_path)
    print(f"final={out_path}", flush=True)
    return out_path


def _debug_dump(data, section):
    """Pretty-print a slice of the hashmap for --debug <section>."""
    section = section.lower()
    print(f"=== DEBUG: {section} ===")
    shown = 0
    for key, metrics in sorted(data.items()):
        if section == "ipc":
            val = metrics.get("ipc", SENTINEL_RUNNING)
            print(f"  {key} -> ipc={val}")
        elif section in ("loads", "miss", "mpki", "byp"):
            field_map = {
                "loads": "loads",
                "miss": "load_miss",
                "mpki": "MPKI",
                "byp": "byp_req",
            }
            fld = field_map[section]
            triple = tuple(metrics.get((lvl, fld), SENTINEL_RUNNING) for lvl in LEVELS_TRIPLE)
            print(f"  {key} -> {fld}{triple}")
        elif section == "cycle_class":
            lpm_keys = [k for k in metrics if isinstance(k, tuple) and k[1] in
                        ("APC", "kappa", "phi", "C-AMAT", "MST", "LPMR_std", "LPMR_byp")]
            print(f"  {key} -> {len(lpm_keys)} LPM fields")
        elif section == "raw":
            print(f"  {key} -> {metrics}")
        else:
            print(f"  (unknown section '{section}', try: ipc|loads|miss|mpki|byp|cycle_class|raw)")
            return
        shown += 1
        if shown >= 50:
            print(f"  ... ({len(data) - 50} more entries suppressed)")
            break


def main():
    args = sys.argv[1:]
    debug_section = None
    n_workers = 8

    if "--debug" in args:
        i = args.index("--debug")
        if i + 1 >= len(args):
            print("usage: aggregator_parallel.py <outputs-dir> [--workers N] [--debug <section>]", file=sys.stderr)
            sys.exit(2)
        debug_section = args[i + 1]
        args = args[:i] + args[i + 2:]

    if "--workers" in args:
        i = args.index("--workers")
        if i + 1 >= len(args):
            print("usage: aggregator_parallel.py <outputs-dir> [--workers N] [--debug <section>]", file=sys.stderr)
            sys.exit(2)
        n_workers = int(args[i + 1])
        args = args[:i] + args[i + 2:]

    if len(args) != 1:
        print("usage: aggregator_parallel.py <outputs-dir> [--workers N] [--debug <section>]", file=sys.stderr)
        sys.exit(2)
    outputs_dir = args[0]
    if not os.path.isdir(outputs_dir):
        print(f"not a dir: {outputs_dir}", file=sys.stderr)
        sys.exit(2)

    input_basename = os.path.basename(os.path.abspath(outputs_dir.rstrip("/")))

    print(f"workers={n_workers} chunksize=256", flush=True)
    if FILTER_INPUT:
        print("WARN: FILTER_INPUT=True — ALLOWED_TRACES/ALLOWED_MODELS whitelist ACTIVE, some logs may be skipped", flush=True)
    data, counts, unparsed_names = collect(outputs_dir, n_workers=n_workers, chunksize=256)
    n_added, n_expected = deduce_missing(data)
    counts["deduced_missing"] = n_added
    counts["expected_matrix"] = n_expected

    if debug_section is not None:
        _debug_dump(data, debug_section)

    out_path = write_xlsx(data, input_basename, n_workers=n_workers)

    print(
        f"logs={counts['logs']} parsed={counts['parsed']} "
        f"ok={counts['ok']} running={counts['running']} errored={counts['errored']} "
        f"deduced_missing={counts['deduced_missing']}/{counts['expected_matrix']} "
        f"out={os.path.basename(out_path) if os.path.isfile(out_path) else out_path}"
    )

    grep_count, files_with_ipc = validate_ipc_count(outputs_dir, counts["ok"])
    if grep_count is not None:
        if grep_count != counts["ok"]:
            print(
                f"WARN: grep found FINAL ROI in {grep_count} files but python "
                f"extracted {counts['ok']}  (delta={grep_count - counts['ok']})"
            )
            py_ok_basenames = set()
            # re-walk and record which logs python classified as ok
            for entry in os.listdir(outputs_dir):
                core_match = CORE_DIR_RE.match(entry)
                if not core_match:
                    continue
                cd = os.path.join(outputs_dir, entry)
                if not os.path.isdir(cd):
                    continue
                for fn in os.listdir(cd):
                    if not fn.endswith(".log"):
                        continue
                    meta = parse_log_filename(fn)
                    if meta is None:
                        continue
                    text = read_log(os.path.join(cd, fn))
                    ipc, sent = extract_ipc(text)
                    if sent == 0:
                        py_ok_basenames.add(os.path.abspath(os.path.join(cd, fn)))
            grep_set = {os.path.abspath(p) for p in files_with_ipc}
            missing = sorted(grep_set - py_ok_basenames)
            for p in missing[:20]:
                print(f"  MISSING: {os.path.basename(p)}")
            if len(missing) > 20:
                print(f"  ... and {len(missing) - 20} more")
        else:
            print(f"OK: grep/python IPC count match ({grep_count})")

    if unparsed_names:
        print(f"WARN: {len(unparsed_names)} log filenames did not match parser")
        for n in unparsed_names[:10]:
            print(f"  UNPARSED: {os.path.basename(n)}")
        if len(unparsed_names) > 10:
            print(f"  ... and {len(unparsed_names) - 10} more")


if __name__ == "__main__":
    main()