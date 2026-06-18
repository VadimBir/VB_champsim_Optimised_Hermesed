#!/usr/bin/env python3
"""
HETERO-OCP weighted-speedup Excel builder (MARS vs SOTA OCP).

Replicates the user's hand-built HETERO methodology with LIVE Excel formulas
(so every result is auditable in the sheet):

  per-core IPC_Alone = nested-IF lookup of the trace name into the 1c base table
  per-core ratio     = IPC_Together / IPC_Alone
  group sum          = SUM(ratio over the 4 cores)
  normalized speedup = group_sum / BASE_group_sum
                       BASE = no-PF LRU 'none' group-sum  (THE SINGLE BASE,
                       used for BOTH the NO_PF and WITH_PF sheets so both
                       conditions live on one comparable scale -> single chart)

Sheets:
  NO_PF   : PF='no'. IPC_Together raw values; everything else formulas.
            none-row group-sum is the base (norm of none == 1.0).
  WITH_PF : per-core AVG over the 4 SOTA prefetchers (bingo_dpc3, spp, berti_c,
            Zion1) via =AVERAGE(); norm divides by the NO_PF none group-sum
            (cross-sheet ref) -> NOT 1.0 for none (captures PF uplift).
  CHART   : summary table (formulas referencing the two sheets) sorted by MARS
            NO_PF speedup ASCENDING, plus ONE embedded scatter chart.

Alone IPC source: core1, filtered to the no-PF base (pf=no, model=no, ocp=off).
"""
import os, re, sys, glob

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------------------------------------------------------- config
BASE_DIR_DEFAULT = "/home/cc/champsim_VB/outputs/DONE_FINAL_HETERO-glc"
OUT_DEFAULT      = "/home/cc/champsim_VB/999_READY_EXCELS/HETERO_OCP_MARS_speedup.xlsx"

SOTA_PF   = ["bingo_dpc3", "spp", "berti_c", "Zion1"]   # 4 prefetchers averaged for WITH_PF
NOPF      = "no"
MARS_BYP  = "4000fix-KappaPhiL1L2"                      # filename token => technique MARS
OCP_SET   = ("hermes", "hmp", "ttp")
TECH_ORDER = ["none", "MARS", "hermes", "hmp", "ttp"]   # row order per mix
BASE_TECH  = "none"                                     # normalization denominator technique
CHART_TECH = ["MARS", "hermes", "hmp", "ttp"]           # plotted series (none==1.0 in NO_PF, omitted)
MARK = {"MARS": "diamond", "hermes": "triangle", "hmp": "square", "ttp": "x"}

IPC_RE      = re.compile(r'FINAL ROI CORE AVG IPC:\s*;([0-9.]+);')
FINISHED_RE = re.compile(r'Finished CPU ;(\d+);.*?AVG IPC: ;([0-9.]+);')
TRACE_RE    = re.compile(r'^trace_(\d+)\s+\S*/([^/]+?)\.champsimtrace', re.M)


# ---------------------------------------------------------------- parsing
def technique_of(byp_triple, ocp):
    l1, l2, l3 = byp_triple
    if l1 == MARS_BYP and ocp == "off":
        return "MARS"
    if byp_triple == ("no", "no", "no"):
        if ocp == "off":
            return "none"
        if ocp in OCP_SET:
            return ocp
    return None


def read_alone(base_dir):
    """core1 -> {trace_name: alone_IPC} for the no-PF base."""
    alone = {}
    cdir = os.path.join(base_dir, "core1")
    for f in glob.glob(os.path.join(cdir, "*..1c..*.log")):
        p = os.path.basename(f)[:-4].split("..")
        if len(p) < 14:
            continue
        pf = (p[3], p[4], p[5]); byp = (p[9], p[10], p[11]); ocp = p[13]
        if pf == ("no", "no", "no") and byp == ("no", "no", "no") and ocp == "off":
            with open(f, "r", errors="replace") as fh:
                m = IPC_RE.search(fh.read())
            if m:
                alone[p[2]] = float(m.group(1))
    return alone


def parse_core4(base_dir, alone):
    """
    data[mix][tech][pf_label] = {traces[4], ipc[4], ratio[4], gs}
    mix = (t0,t1,t2,t3). Returns (data, warnings).
    """
    data = {}
    warn = []
    cdir = os.path.join(base_dir, "core4")
    for f in sorted(glob.glob(os.path.join(cdir, "glc..*.log"))):
        p = os.path.basename(f)[:-4].split("..")
        if len(p) < 14:
            continue
        traces = p[2].split("__")
        if len(traces) != 4:
            warn.append("NOT_4_TRACES: " + f)
            continue
        pf_label = p[4]                       # l1/l3 always 'no'; l2 carries the pf
        tech = technique_of((p[9], p[10], p[11]), p[13])
        if tech is None:
            warn.append("UNKNOWN_TECH: " + f)
            continue
        with open(f, "r", errors="replace") as fh:
            txt = fh.read()
        if not IPC_RE.search(txt):
            warn.append("NO_FINAL_ROI: " + f)
            continue
        cpu_trace = {int(m.group(1)): m.group(2) for m in TRACE_RE.finditer(txt)}
        cpu_ipc = {int(m.group(1)): float(m.group(2)) for m in FINISHED_RE.finditer(txt)}
        if len(cpu_ipc) < 4:
            warn.append("MISSING_CORE_IPC(%d): %s" % (len(cpu_ipc), f))
            continue
        ipc = []; ratio = []; tr_names = []; ok = True
        for c in range(4):
            tr = cpu_trace.get(c, traces[c])
            iv = cpu_ipc.get(c)
            av = alone.get(tr)
            if iv is None or av is None:
                warn.append("NO_ALONE[%s] or NO_IPC core%d: %s" % (tr, c, f))
                ok = False; break
            tr_names.append(tr); ipc.append(iv); ratio.append(iv / av)
        if not ok:
            continue
        mix = tuple(traces)
        rec = {"traces": tr_names, "ipc": ipc, "ratio": ratio, "gs": sum(ratio)}
        data.setdefault(mix, {}).setdefault(tech, {})[pf_label] = rec
    return data, warn


# ---------------------------------------------------------------- compute (python mirror, for sort/values)
def compute_nopf(data):
    out = {}
    for mix, td in data.items():
        base = td.get(BASE_TECH, {}).get(NOPF)
        if not base:
            continue
        row = {}
        for tech in TECH_ORDER:
            rec = td.get(tech, {}).get(NOPF)
            if not rec:
                continue
            row[tech] = {"traces": rec["traces"], "ipc": rec["ipc"],
                         "ratio": rec["ratio"], "gs": rec["gs"],
                         "norm": rec["gs"] / base["gs"]}
        out[mix] = row
    return out


def compute_withpf(data, nopf):
    """avg over 4 SOTA PF; norm = withpf_gs / NO-PF none gs (single base)."""
    out = {}
    for mix, td in data.items():
        if mix not in nopf or BASE_TECH not in nopf[mix]:
            continue
        nopf_base_gs = nopf[mix][BASE_TECH]["gs"]

        def avg_rec(tech):
            recs = []
            for pf in SOTA_PF:
                rec = td.get(tech, {}).get(pf)
                if rec:
                    recs.append(rec)
            if len(recs) != len(SOTA_PF):
                return None
            pf_ipc = {pf: td[tech][pf]["ipc"] for pf in SOTA_PF}
            avg_ipc = [sum(r["ipc"][c] for r in recs) / len(recs) for c in range(4)]
            alone = [recs[0]["ipc"][c] / recs[0]["ratio"][c] for c in range(4)]
            ratio = [avg_ipc[c] / alone[c] for c in range(4)]
            return {"traces": recs[0]["traces"], "pf_ipc": pf_ipc, "avg_ipc": avg_ipc,
                    "alone": alone, "ratio": ratio, "gs": sum(ratio)}
        row = {}
        for tech in TECH_ORDER:
            rec = avg_rec(tech)
            if not rec:
                continue
            rec["norm"] = rec["gs"] / nopf_base_gs
            row[tech] = rec
        if row:
            out[mix] = row
    return out


# ---------------------------------------------------------------- excel helpers
H_FONT = Font(bold=True)
H_FILL = PatternFill("solid", fgColor="DDEBF7")
CEN    = Alignment(horizontal="center")


def cr(col, row):
    return "%s%d" % (get_column_letter(col), row)


def _hdr(ws, r, c, text):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CEN
    return cell


def write_alone_table(ws, alone_order):
    """alone_order: list of (trace, brow). IPC at B{brow}, name at A{brow}. rows 1..N."""
    ws.cell(row=1, column=7, value="1c Alone IPC base (no PF):  A=trace  B=IPC").font = H_FONT
    for tr, br in alone_order:
        ws.cell(row=br, column=1, value=tr)
        ws.cell(row=br, column=2, value=alone[tr])


def alone_if_formula(name_ref, alone_order):
    s = ""
    for tr, br in alone_order:
        s += 'IF(%s="%s",$B$%d,' % (name_ref, tr, br)
    return "=" + s + '""' + (")" * len(alone_order))


# ---------------------------------------------------------------- NO_PF sheet
# column plan (matches user's hand layout: names@C, together@G, alone@P, ratio@Y, gs@AH)
N_TECH = 1            # A
N_NAME = 3            # C..F  (3..6)
N_TOG  = 7            # G..J  (7..10)
N_ALN  = 16          # P..S  (16..19)
N_RAT  = 25          # Y..AB (25..28)
N_GS   = 34          # AH
N_NORM = 36          # AJ


def write_nopf_sheet(wb, nopf, alone_order, trow):
    ws = wb.create_sheet("NO_PF")
    write_alone_table(ws, alone_order)
    top = trow; sub = trow + 1

    _hdr(ws, top, N_TOG, "IPC Together")
    _hdr(ws, top, N_ALN, "IPC_Alone (1c base, no PF)")
    _hdr(ws, top, N_RAT, "IPC Together / IPC_Alone")
    _hdr(ws, top, N_GS,  "Group SUM Together/Alone")
    _hdr(ws, top, N_NORM, "Relative weighted speedup (vs no-PF none)")
    for j, lab in enumerate(["C1", "C2", "C3", "C4"]):
        _hdr(ws, sub, N_NAME + j, lab)
        _hdr(ws, sub, N_TOG + j, lab)
        _hdr(ws, sub, N_ALN + j, lab)
        _hdr(ws, sub, N_RAT + j, lab)

    r = sub + 1
    base_row = {}        # mix -> none gs row (AH)
    norm_row = {}        # (mix,tech) -> norm row (AJ)
    for mix in sorted(nopf):
        row = nopf[mix]
        # none first so base_row is known
        mix_base = None
        for tech in TECH_ORDER:
            if tech not in row:
                continue
            d = row[tech]
            ws.cell(row=r, column=N_TECH, value=tech)
            for c in range(4):
                ws.cell(row=r, column=N_NAME + c, value=d["traces"][c])
                ws.cell(row=r, column=N_TOG + c, value=d["ipc"][c])           # raw measured
                ws.cell(row=r, column=N_ALN + c,
                        value=alone_if_formula(cr(N_NAME + c, r), alone_order))
                ws.cell(row=r, column=N_RAT + c,
                        value="=%s/%s" % (cr(N_TOG + c, r), cr(N_ALN + c, r)))
            ws.cell(row=r, column=N_GS,
                    value="=SUM(%s:%s)" % (cr(N_RAT, r), cr(N_RAT + 3, r)))
            if tech == BASE_TECH:
                mix_base = r
                base_row[mix] = r
            ws.cell(row=r, column=N_NORM,
                    value="=%s/%s" % (cr(N_GS, r), cr(N_GS, mix_base)))
            norm_row[(mix, tech)] = r
            r += 1
    write_scatter_and_chart(ws, nopf, norm_row, N_NORM, scol=40,
                            hrow=sub, drow=sub + 1,
                            title="HETERO Weighted Speedup — NO Prefetch (vs no-PF-LRU base)")
    return ws, base_row, norm_row


# ---------------------------------------------------------------- WITH_PF sheet
W_TECH  = 1           # A
W_NAME  = 3           # C..F (3..6)
W_CORE0 = 8           # H ; per-core block = 5 cols (4 PF + AVG)
W_PER   = 5
W_ALN   = 29         # AC..AF (29..32)
W_RAT   = 34         # AH..AK (34..37)
W_GS    = 39         # AM
W_NORM  = 41         # AO


def write_withpf_sheet(wb, withpf, alone_order, trow, nopf_base_row):
    ws = wb.create_sheet("WITH_PF")
    write_alone_table(ws, alone_order)
    top = trow; sub = trow + 1

    for c in range(4):
        b = W_CORE0 + c * W_PER
        _hdr(ws, top, b, "Core%d IPC (4 SOTA PF + AVG)" % c)
        for j, pf in enumerate(SOTA_PF):
            _hdr(ws, sub, b + j, pf)
        _hdr(ws, sub, b + len(SOTA_PF), "AVG")
    _hdr(ws, top, W_ALN, "IPC_Alone (1c base, no PF)")
    _hdr(ws, top, W_RAT, "AVG IPC / IPC_Alone")
    _hdr(ws, top, W_GS,  "Group SUM Together/Alone")
    _hdr(ws, top, W_NORM, "Relative speedup (vs NO_PF none, single base)")
    for j, lab in enumerate(["C1", "C2", "C3", "C4"]):
        _hdr(ws, sub, W_NAME + j, lab)
        _hdr(ws, sub, W_ALN + j, lab)
        _hdr(ws, sub, W_RAT + j, lab)

    r = sub + 1
    norm_row = {}
    for mix in sorted(withpf):
        if mix not in nopf_base_row:
            continue
        nb = nopf_base_row[mix]                       # NO_PF none gs row (col N_GS=AH)
        row = withpf[mix]
        for tech in TECH_ORDER:
            if tech not in row:
                continue
            d = row[tech]
            ws.cell(row=r, column=W_TECH, value=tech)
            for c in range(4):
                ws.cell(row=r, column=W_NAME + c, value=d["traces"][c])
                b = W_CORE0 + c * W_PER
                for j, pf in enumerate(SOTA_PF):
                    ws.cell(row=r, column=b + j, value=d["pf_ipc"][pf][c])   # raw measured
                ws.cell(row=r, column=b + len(SOTA_PF),
                        value="=AVERAGE(%s:%s)" % (cr(b, r), cr(b + 3, r)))
                ws.cell(row=r, column=W_ALN + c,
                        value=alone_if_formula(cr(W_NAME + c, r), alone_order))
                ws.cell(row=r, column=W_RAT + c,
                        value="=%s/%s" % (cr(b + len(SOTA_PF), r), cr(W_ALN + c, r)))
            ws.cell(row=r, column=W_GS,
                    value="=SUM(%s:%s)" % (cr(W_RAT, r), cr(W_RAT + 3, r)))
            ws.cell(row=r, column=W_NORM,
                    value="=%s/NO_PF!%s" % (cr(W_GS, r), cr(N_GS, nb)))
            norm_row[(mix, tech)] = r
            r += 1
    write_scatter_and_chart(ws, withpf, norm_row, W_NORM, scol=45,
                            hrow=sub, drow=sub + 1,
                            title="HETERO Weighted Speedup — WITH Prefetch avg4 (vs no-PF-LRU base)")
    return ws, norm_row


# ---------------------------------------------------------------- per-sheet scatter block + chart
def write_scatter_and_chart(ws, comp, norm_row, norm_col, scol, hrow, drow, title):
    """
    Sorted (by MARS norm asc) scatter table on the SAME sheet, cells are FORMULAS
    referencing the data block's norm column, plus an embedded 4-series chart.
    hrow = header row (aligned with data sub-header), drow = first data row.
    """
    order = sorted((m for m in comp if "MARS" in comp[m]),
                   key=lambda m: comp[m]["MARS"]["norm"])
    _hdr(ws, hrow, scol, "rank")
    for j, tech in enumerate(CHART_TECH):
        _hdr(ws, hrow, scol + 1 + j, tech)
    for i, mix in enumerate(order):
        rr = drow + i
        ws.cell(row=rr, column=scol, value=i + 1)
        for j, tech in enumerate(CHART_TECH):
            src = norm_row.get((mix, tech))
            if src is None:
                continue
            ws.cell(row=rr, column=scol + 1 + j, value="=%s" % cr(norm_col, src))
    last = drow + len(order) - 1

    ch = ScatterChart()
    ch.title = title
    ch.style = 13
    ch.x_axis.title = "Mix (sorted by MARS speedup ascending)"
    ch.y_axis.title = "Normalized Weighted Speedup"
    ch.x_axis.delete = False; ch.y_axis.delete = False
    ch.height = 11; ch.width = 26
    xref = Reference(ws, min_col=scol, min_row=drow, max_row=last)
    for j, tech in enumerate(CHART_TECH):
        yref = Reference(ws, min_col=scol + 1 + j, min_row=hrow, max_row=last)
        s = Series(yref, xref, title_from_data=True)
        s.marker = Marker(symbol=MARK[tech], size=4)
        s.graphicalProperties.line.noFill = True
        ch.series.append(s)
    ws.add_chart(ch, cr(scol, last + 3))


# ---------------------------------------------------------------- main
def main():
    global alone
    base_dir = sys.argv[1] if len(sys.argv) > 1 else BASE_DIR_DEFAULT
    out = sys.argv[2] if len(sys.argv) > 2 else OUT_DEFAULT

    alone = read_alone(base_dir)
    print("alone traces:", len(alone))
    data, warn = parse_core4(base_dir, alone)
    print("mixes:", len(data), "warnings:", len(warn))

    nopf = compute_nopf(data)
    withpf = compute_withpf(data, nopf)
    print("nopf mixes:", len(nopf), "withpf mixes:", len(withpf))

    alone_order = [(tr, i + 1) for i, tr in enumerate(sorted(alone))]
    trow = max(18, len(alone) + 4)

    wb = Workbook(); wb.remove(wb.active)
    _, nopf_base_row, nopf_norm_row = write_nopf_sheet(wb, nopf, alone_order, trow)
    _, withpf_norm_row = write_withpf_sheet(wb, withpf, alone_order, trow, nopf_base_row)

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print("WROTE:", out)

    if warn:
        wf = out + "_WARNINGS.txt"
        with open(wf, "w") as f:
            f.write("\n".join(warn) + "\n")
        print("WARNINGS:", len(warn), "->", wf)


if __name__ == "__main__":
    main()
