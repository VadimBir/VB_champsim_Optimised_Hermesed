#!/usr/bin/env python3
"""Generate FIGURE_TEMPLATES.xlsx — style/layout templates for the paper figures.

Three chart archetypes (cover all 6 planned figures), each with DUMMY data so the
style/size/colors can be judged + hand-tuned before wiring real data:
  T1_perPF     — wide multi-core grouped bar, PF gradient (F1)
  T2_technique — wide multi-core grouped bar, white Base -> grays OCP -> black MARS (F2/F4/F4b/F5)
  T3_perRepl   — per-repl small-multiple panels, 4.3 x 2.1 cm each (F3)

Opens in Excel AND LibreOffice Calc (.xlsx). Reuses chart_figs styling helpers so
template look == final pipeline look.
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import chart_figs as cf
from openpyxl import Workbook
from openpyxl.chart import Reference

OUT = "/home/cc/champsim_VB/999_READY_EXCELS/FIGURE_TEMPLATES.xlsx"

CORES = ["1 Core", "4 Core", "8 Core", "16 Core"]


def _write_grid(ws, hdr_row, headers, rows):
    ws.cell(row=hdr_row, column=1, value=headers[0])
    for j, h in enumerate(headers[1:]):
        ws.cell(row=hdr_row, column=2 + j, value=h)
    for i, (label, vals) in enumerate(rows):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=label)
        for j, v in enumerate(vals):
            ws.cell(row=r, column=2 + j, value=v)


def t1_perPF(wb):
    ws = wb.create_sheet("T1_perPF")
    ws["A1"] = "TEMPLATE T1 (F1): per-PF speedup, X=cores, light->black PF gradient. DUMMY data."
    series = ["SPP", "Zion", "Bingo", "Berti"]
    rows = [
        ("1 Core", [1.23, 1.34, 1.31, 1.36]),
        ("4 Core", [1.24, 1.41, 1.35, 1.37]),
        ("8 Core", [1.19, 1.35, 1.27, 1.29]),
        ("16 Core", [1.10, 1.20, 1.18, 1.16]),
    ]
    hdr = 3
    _write_grid(ws, hdr, ["Cores"] + series, rows)
    cat = Reference(ws, min_col=1, min_row=hdr + 1, max_row=hdr + len(rows))
    dat = Reference(ws, min_col=2, max_col=1 + len(series), min_row=hdr, max_row=hdr + len(rows))
    cf._grouped_bar(ws, "Normalized Speedup", cat, dat, "H3",
                    y_min=1.0, colors=cf._gray_gradient(len(series)))


def t2_technique(wb):
    ws = wb.create_sheet("T2_technique")
    ws["A1"] = "TEMPLATE T2 (F2/F4/F4b/F5): white Base -> grays OCP -> BLACK MARS (rightmost). DUMMY."
    series = ["Base", "HERMES", "HMP", "TTP", "MARS"]
    rows = [
        ("1 Core", [1.00, 1.03, 1.02, 1.04, 1.13]),
        ("4 Core", [1.00, 1.05, 1.04, 1.06, 1.18]),
        ("8 Core", [1.00, 1.04, 1.03, 1.05, 1.15]),
        ("16 Core", [1.00, 1.02, 1.02, 1.03, 1.11]),
    ]
    hdr = 3
    _write_grid(ws, hdr, ["Cores"] + series, rows)
    cat = Reference(ws, min_col=1, min_row=hdr + 1, max_row=hdr + len(rows))
    dat = Reference(ws, min_col=2, max_col=1 + len(series), min_row=hdr, max_row=hdr + len(rows))
    # gradient gives white(Base) -> grays(OCP) -> black(MARS, rightmost) by series order
    cf._grouped_bar(ws, "Speedup vs LRU no-pref base", cat, dat, "H3",
                    y_min=1.0, colors=cf._gray_gradient(len(series)))


def t3_perRepl(wb):
    ws = wb.create_sheet("T3_perRepl")
    ws["A1"] = "TEMPLATE T3 (F3): per-repl small-multiple panels (4.3 x 2.1 cm each). DUMMY."
    repls = ["care", "hawkeye", "ship++"]
    series = ["no-pref", "with-pref", "MARS"]
    anchors = ["H3", "H15", "H27"]
    for k, repl in enumerate(repls):
        hdr = 3 + k * 10
        ws.cell(row=hdr - 1, column=1, value=f"repl={repl}")
        rows = [
            ("1 Core", [1.00, 1.05 + 0.01 * k, 1.12 + 0.01 * k]),
            ("4 Core", [1.00, 1.07 + 0.01 * k, 1.16 + 0.01 * k]),
            ("8 Core", [1.00, 1.06 + 0.01 * k, 1.14 + 0.01 * k]),
            ("16 Core", [1.00, 1.03 + 0.01 * k, 1.10 + 0.01 * k]),
        ]
        _write_grid(ws, hdr, ["Cores"] + series, rows)
        cat = Reference(ws, min_col=1, min_row=hdr + 1, max_row=hdr + len(rows))
        dat = Reference(ws, min_col=2, max_col=1 + len(series), min_row=hdr, max_row=hdr + len(rows))
        cf._grouped_bar(ws, repl, cat, dat, anchors[k],
                        y_min=1.0, colors=cf._gray_gradient(len(series)),
                        width_cm=4.3, height_cm=2.1)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    t1_perPF(wb)
    t2_technique(wb)
    t3_perRepl(wb)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
