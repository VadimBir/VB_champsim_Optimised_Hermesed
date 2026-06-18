#!/usr/bin/env python3
"""Test 1: Speedup formula with new base logic.

Layout (6 data cols, within bounds):
  Row labels col A, then:
  Trace1: Base(B), Hermes(C), Fwd(D)  |  Trace2: Base(E), Hermes(F), Fwd(G)

Rows:
  R1: trace header (merged)
  R2: sub-col headers
  R3: no-pf IPC values (base row)
  R4: berti_c IPC values (PF row)
  R5: speedup for no-pf (should be 1.0 for Base, ratio for others)
  R6: speedup for berti_c (each subcol / no-pf Base)
  R7: GM across traces for each sub-col
"""
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Speedup Test"

# --- Config ---
traces = ["605.mcf", "429.bzip"]
subcols = ["Base", "Hermes", "Fwd"]
n_sub = len(subcols)
pf_rows = ["no-pf", "berti_c"]

# Sample IPC values: [trace][subcol] for each pf_row
ipc = {
    "no-pf":    {"605.mcf": [1.20, 1.18, 1.15], "429.bzip": [2.00, 1.95, 1.90]},
    "berti_c":  {"605.mcf": [1.50, 1.48, 1.55], "429.bzip": [2.40, 2.35, 2.50]},
}

# --- Row 1: Trace headers (merged) ---
data_start_col = 2  # col B
for ti, trace in enumerate(traces):
    c1 = data_start_col + ti * n_sub
    c2 = c1 + n_sub - 1
    ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
    ws.cell(row=1, column=c1, value=trace)

# --- Row 2: Sub-col headers ---
ws.cell(row=2, column=1, value="PF")
for ti in range(len(traces)):
    for si, sc in enumerate(subcols):
        col = data_start_col + ti * n_sub + si
        ws.cell(row=2, column=col, value=sc)

# --- Rows 3-4: IPC values ---
ipc_start_row = 3
for pi, pf in enumerate(pf_rows):
    row = ipc_start_row + pi
    ws.cell(row=row, column=1, value=pf)
    for ti, trace in enumerate(traces):
        for si, val in enumerate(ipc[pf][trace]):
            col = data_start_col + ti * n_sub + si
            ws.cell(row=row, column=col, value=val)

# --- Rows 5-6: Speedup formulas ---
# NEW BASE LOGIC:
#   speedup = this_cell / no-pf_Base_cell_for_same_trace
#   no-pf Base col for trace ti = data_start_col + ti * n_sub (the first subcol)
speedup_start_row = 6
ws.cell(row=5, column=1, value="--- Speedup ---")
for pi, pf in enumerate(pf_rows):
    row = speedup_start_row + pi
    ws.cell(row=row, column=1, value=f"SU:{pf}")
    ipc_row = ipc_start_row + pi
    base_ipc_row = ipc_start_row  # no-pf row is always the base denominator

    for ti in range(len(traces)):
        base_col = data_start_col + ti * n_sub  # Base subcol for this trace
        base_cell = f"${get_column_letter(base_col)}${base_ipc_row}"  # no-pf Base

        for si in range(n_sub):
            col = data_start_col + ti * n_sub + si
            num_cell = f"{get_column_letter(col)}{ipc_row}"
            formula = f"={num_cell}/{base_cell}"
            ws.cell(row=row, column=col, value=formula)
            print(f"  [{pf}][trace{ti}][sub{si}] {get_column_letter(col)}{row} = {formula}")

# --- Row 8: GM across traces (non-contiguous) ---
gm_row = 9
ws.cell(row=8, column=1, value="--- GM ---")
for pi, pf in enumerate(pf_rows):
    row = gm_row + pi
    ws.cell(row=row, column=1, value=f"GM:{pf}")
    su_row = speedup_start_row + pi

    for si in range(n_sub):
        # Collect cell refs for this subcol across all traces (non-contiguous)
        refs = []
        for ti in range(len(traces)):
            col = data_start_col + ti * n_sub + si
            refs.append(f"{get_column_letter(col)}{su_row}")

        n = len(refs)
        # GM = PRODUCT(cells)^(1/n) via GEOMEAN or manual
        # openpyxl: GEOMEAN works if cells non-contiguous passed as args
        gm_formula = f"=GEOMEAN({','.join(refs)})"
        out_col = data_start_col + si  # put GM in first trace's subcol positions
        ws.cell(row=row, column=out_col, value=gm_formula)
        print(f"  GM[{pf}][sub{si}] {get_column_letter(out_col)}{row} = {gm_formula}")

outpath = "/home/cc/champsim_VB/PyOutProcessScript/test_snippets/test_speedup.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")

# --- Verify by reading back ---
print("\n--- Readback ---")
wb2 = openpyxl.load_workbook(outpath)
ws2 = wb2.active
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=ws2.max_column, values_only=False):
    vals = [f"{c.value}" if c.value is not None else "" for c in row]
    print("  ".join(f"{v:>16}" for v in vals))
