#!/usr/bin/env python3
"""Test 3: Group header with merged cells.

Layout: 2 traces x 3 sub-cols = 6 data cols (within bounds).
5 gap cols between trace groups.
Row 1: Trace name merged across 3 sub-cols.
Row 2: Sub-col labels (Base, Hermes, Fwd).
Row 3+: sample data rows.
"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Header Test"

traces = ["605.mcf_s", "429.bzip2"]
subcols = ["Base", "Hermes", "Fwd"]
n_sub = len(subcols)
gap_cols = 5
label_col = 1  # col A for row labels

# Calculate trace start columns
# Trace 0 starts at col 2 (B)
# Trace 1 starts at col 2 + n_sub + gap_cols
trace_starts = []
for ti in range(len(traces)):
    start = label_col + 1 + ti * (n_sub + gap_cols)
    trace_starts.append(start)

print("Trace start columns:")
for ti, ts in enumerate(trace_starts):
    print(f"  {traces[ti]}: col {ts} ({get_column_letter(ts)})")

# --- Styling ---
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF", size=11)
sub_fill = PatternFill("solid", fgColor="D9E2F3")
sub_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# --- Row 1: Merged trace headers ---
for ti, trace in enumerate(traces):
    c1 = trace_starts[ti]
    c2 = c1 + n_sub - 1
    ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
    cell = ws.cell(row=1, column=c1, value=trace)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
    # Apply border to merged cells too
    for c in range(c1 + 1, c2 + 1):
        ws.cell(row=1, column=c).border = thin_border

# --- Row 2: Sub-col headers ---
ws.cell(row=2, column=1, value="Prefetcher")
for ti in range(len(traces)):
    for si, sc in enumerate(subcols):
        col = trace_starts[ti] + si
        cell = ws.cell(row=2, column=col, value=sc)
        cell.fill = sub_fill
        cell.font = sub_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

# --- Row 3+: Sample data ---
sample_pfs = ["no-pf", "berti_c"]
for pi, pf in enumerate(sample_pfs):
    row = 3 + pi
    ws.cell(row=row, column=1, value=pf)
    for ti in range(len(traces)):
        for si in range(n_sub):
            col = trace_starts[ti] + si
            val = 1.0 + ti * 0.5 + si * 0.05 + pi * 0.3
            ws.cell(row=row, column=col, value=round(val, 3))
            ws.cell(row=row, column=col).border = thin_border

# --- Gap cols: verify they're empty ---
if len(traces) > 1:
    gap_start = trace_starts[0] + n_sub
    gap_end = trace_starts[1] - 1
    print(f"Gap columns: {get_column_letter(gap_start)} to {get_column_letter(gap_end)} ({gap_end - gap_start + 1} cols)")
    for gc in range(gap_start, gap_end + 1):
        for r in range(1, 5):
            assert ws.cell(row=r, column=gc).value is None, f"Gap col {gc} row {r} not empty!"
    print("Gap cols verified empty.")

# --- Column widths ---
ws.column_dimensions['A'].width = 14
for ti in range(len(traces)):
    for si in range(n_sub):
        col_letter = get_column_letter(trace_starts[ti] + si)
        ws.column_dimensions[col_letter].width = 12

outpath = "/home/cc/champsim_VB/PyOutProcessScript/test_snippets/test_header.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")

# --- Readback ---
print("\n--- Readback (non-empty cells) ---")
wb2 = openpyxl.load_workbook(outpath)
ws2 = wb2.active
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=ws2.max_column):
    parts = []
    for c in row:
        if c.value is not None:
            parts.append(f"{get_column_letter(c.column)}{c.row}={c.value}")
    if parts:
        print("  " + "  |  ".join(parts))

print("\nMerged ranges:", list(ws2.merged_cells.ranges))
