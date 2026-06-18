#!/usr/bin/env python3
"""ChampSim results aggregator — Checkpoint 1: bare-min IPC only.

Usage:
    python3 aggregator.py <outputs-dir>

Walks <outputs-dir>/core*/ subdirs for .log files. Extracts FINAL ROI CORE
AVG IPC per run. Writes xlsx to ../999_READY_EXCELS/.

See plan: ~/.claude/plans/cryptic-plotting-noodle.md
"""

import atexit
import os
import re
import signal
import subprocess
import sys
import time
from collections import defaultdict

# CP9: global failure tracker — appended during collect(), surfaced by
# _print_failure_report() via atexit + SIGINT/SIGTERM trap so even an
# interrupted run prints the list of bad logs before exiting.
FAILED_LOGS = []  # list of (path, reason)


def _print_failure_report():
    if not FAILED_LOGS:
        return
    print("=== WARNING: failed/incomplete logs ===", file=sys.stderr)
    for p, reason in FAILED_LOGS[:200]:
        print(f"  {os.path.basename(p)}  reason: {reason}", file=sys.stderr)
    if len(FAILED_LOGS) > 200:
        print(f"  ... and {len(FAILED_LOGS) - 200} more", file=sys.stderr)
    print(f"  total: {len(FAILED_LOGS)} failed/incomplete", file=sys.stderr)


def _signal_trap(signum, _frame):
    _print_failure_report()
    sys.exit(128 + signum)


atexit.register(_print_failure_report)
signal.signal(signal.SIGINT, _signal_trap)
signal.signal(signal.SIGTERM, _signal_trap)

from openpyxl import Workbook

from stat_patterns import (
    STAT_PATTERNS,
    ERROR_MARKERS,
    PER_LEVEL_FIELD_RE,
    SUM_FIELDS,
    AVG_FIELDS,
    CP5_FAMILIES,
    LEVELS_TRIPLE,
)

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

PER_LEVEL_RE_C = re.compile(PER_LEVEL_FIELD_RE)

# cd to script dir so relative paths work
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(SCRIPT_DIR)

OUTPUT_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "999_READY_EXCELS"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Filename format (confirmed from real samples):
#   L1-<l1pf>-L2-<l2pf>-L3-<llcpf>-<TRACE>-<L1m>_<L2m>_<LLCm>.log
# TRACE is of the form 403.gcc-17B or 603.bwaves_s-2931B etc., ends in -<N>B
FILENAME_RE = re.compile(
    r"^(?:\d+-)?L1-(?P<l1pf>.+?)-L2-(?P<l2pf>.+?)-L3-(?P<llcpf>.+?)-"
    r"(?P<trace>.+?-\d+[BM](?:_\d+[BM])?)-"
    r"(?P<modelstr>.+)\.log$"
)

CORE_DIR_RE = re.compile(r"^core(\d+)$")

SENTINEL_RUNNING = -1
SENTINEL_ERRORED = -2


def split_model_triplet(modelstr):
    """Model segment is <L1m>_<L2m>_<LLCm> with homogeneous repeated tokens.

    Strategy: if the string splits into 3 equal parts on an underscore
    boundary, return those. Otherwise return the whole string 3 times
    (single-model spec).
    """
    # Try splitting by exact-thirds underscore partition.
    # Homogeneous: "X_X_X" where X may itself contain underscores.
    # We search for the split where first third == second == third.
    parts = modelstr.split("_")
    n = len(parts)
    if n % 3 == 0:
        third = n // 3
        a = "_".join(parts[:third])
        b = "_".join(parts[third : 2 * third])
        c = "_".join(parts[2 * third :])
        if a == b == c:
            return a, b, c
    # Fallback: whole string treated as single-model (unknown shape)
    return modelstr, modelstr, modelstr


def parse_log_filename(fname):
    m = FILENAME_RE.match(fname)
    if not m:
        return None
    d = m.groupdict()
    l1m, l2m, llcm = split_model_triplet(d["modelstr"])
    return {
        "l1pf": d["l1pf"],
        "l2pf": d["l2pf"],
        "llcpf": d["llcpf"],
        "trace": d["trace"],
        "l1m": l1m,
        "l2m": l2m,
        "llcm": llcm,
    }


def read_log(path):
    try:
        with open(path, "r", errors="replace") as f:
            return f.read()
    except OSError:
        return None


ERROR_MARKER_RE = re.compile(
    r"(?:^|[\s:/])(?:assert(?:ion)?\s+failed|Segmentation\s+fault|SIGSEGV|"
    r"Aborted|Killed|deadlock)(?:[\s:]|$)",
    re.MULTILINE,
)


def has_error_marker(text):
    # RULE 0: avoid substring false positives like 'assertion_rate'.
    # Require anchored markers with boundary characters.
    return bool(ERROR_MARKER_RE.search(text))


ARCH_RE = re.compile(r"\[RUN\]\s+arch\s+base:\s*(\S+)")


def extract_arch(text, dir_fallback=None):
    # RULE 0: never fabricate. If arch line absent, use the outputs-dir
    # basename (e.g. 'out32', 'out33') — each outXX folder corresponds to
    # one arch run, so the dir name is the authoritative label when the
    # in-log line is missing/truncated.
    if text is not None:
        m = ARCH_RE.search(text)
        if m:
            return m.group(1)
    return dir_fallback if dir_fallback else "unknown"


def extract_ipc(text):
    """Return (value, sentinel_used). sentinel_used in {0,-1,-2}."""
    if text is None:
        return SENTINEL_ERRORED, -2
    regex, _scope = STAT_PATTERNS["ipc_final_avg"]
    m = re.search(regex, text)
    if m:
        try:
            return float(m.group(1)), 0
        except ValueError:
            pass
    # no value found — classify
    if has_error_marker(text):
        return SENTINEL_ERRORED, -2
    return SENTINEL_RUNNING, -1


def extract_per_level_fields(text):
    """Parse all `Core_;N; LVL;_;field; val;` lines.

    Returns {(lvl, field): aggregated_scalar}. Aggregation:
      - SUM_FIELDS → sum across cores
      - AVG_FIELDS → average across cores (skipping nan/inf)
      - else → average (default)
    Missing field → not in dict (caller substitutes SENTINEL_RUNNING).
    """
    if text is None:
        return {}
    # (lvl, field) -> list of per-core floats
    raw = defaultdict(list)
    for m in PER_LEVEL_RE_C.finditer(text):
        _core, lvl, field, val_s = m.group(1), m.group(2), m.group(3), m.group(4)
        if val_s in ("nan", "inf", "-nan", "-inf"):
            continue
        try:
            v = float(val_s)
        except ValueError:
            continue
        raw[(lvl, field)].append(v)
    out = {}
    for (lvl, field), vals in raw.items():
        if not vals:
            continue
        if field in SUM_FIELDS:
            out[(lvl, field)] = sum(vals)
        elif field in AVG_FIELDS:
            out[(lvl, field)] = sum(vals) / len(vals)
        else:
            out[(lvl, field)] = sum(vals) / len(vals)
    return out


LPM_HEADER_RE = re.compile(r"=== LPM Cycle Classification\s+CPU\s+(\d+)\s+\(ROI\) ===")
LPM_COLS = (
    "h", "m", "x", "e",
    "omega", "mu", "kappa", "phi",
    "LPMR_std", "LPMR_byp",
    "C-AMAT", "APC", "MST", "CPA",
)
LPM_LEVELS = ("ITLB", "DTLB", "STLB", "L1I", "L1D", "L2C", "LLC", "DRAM")


def extract_lpm_cycle_class(text):
    """Parse all `=== LPM Cycle Classification CPU N (ROI) ===` blocks.

    Returns {(lvl, 'LPM_'+col): aggregated_scalar} averaged across cores.
    Values of 'n/a' / 'nan' / 'inf' are skipped.
    """
    if text is None:
        return {}
    raw = defaultdict(list)
    lines = text.splitlines()
    i = 0
    n = len(lines)
    while i < n:
        if LPM_HEADER_RE.search(lines[i]):
            # skip header + column-label line
            i += 2
            while i < n:
                line = lines[i].strip()
                if not line:
                    break
                # strip | separators
                parts = re.split(r"\s*\|\s*|\s+", line)
                parts = [p for p in parts if p]
                if not parts or parts[0] not in LPM_LEVELS:
                    break
                lvl = parts[0]
                vals = parts[1:]
                # Expect 14 columns. Tolerate short rows.
                for c_idx, col_name in enumerate(LPM_COLS):
                    if c_idx >= len(vals):
                        break
                    v_s = vals[c_idx]
                    if v_s in ("n/a", "nan", "-nan", "inf", "-inf"):
                        continue
                    try:
                        v = float(v_s)
                    except ValueError:
                        continue
                    raw[(lvl, "LPM_" + col_name)].append(v)
                i += 1
            continue
        i += 1
    out = {}
    for k, vals in raw.items():
        if vals:
            out[k] = sum(vals) / len(vals)
    return out


def collect(outputs_dir):
    outputs_dir = os.path.abspath(outputs_dir)
    dir_base = os.path.basename(outputs_dir.rstrip("/"))
    # key = (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm) -> {metric: scalar}
    # metric names: "ipc", or (lvl, field) tuples
    data = {}
    counts = {"logs": 0, "parsed": 0, "ok": 0, "running": 0, "errored": 0}
    unparsed_names = []

    for entry in sorted(os.listdir(outputs_dir)):
        core_match = CORE_DIR_RE.match(entry)
        if not core_match:
            continue
        cores = int(core_match.group(1))
        core_dir = os.path.join(outputs_dir, entry)
        if not os.path.isdir(core_dir):
            continue

        for fname in sorted(os.listdir(core_dir)):
            if not fname.endswith(".log"):
                continue
            counts["logs"] += 1
            meta = parse_log_filename(fname)
            if meta is None:
                unparsed_names.append(f"core{cores}/{fname}")
                continue
            counts["parsed"] += 1
            fpath = os.path.join(core_dir, fname)
            text = read_log(fpath)
            arch = extract_arch(text, dir_fallback=dir_base)
            ipc, sentinel = extract_ipc(text)
            if sentinel == 0:
                counts["ok"] += 1
            elif sentinel == -1:
                counts["running"] += 1
                FAILED_LOGS.append((fpath, "incomplete (no FINAL ROI, no error marker)"))
            else:
                counts["errored"] += 1
                marker = "error marker found"
                for em in ERROR_MARKERS:
                    if re.search(rf"\b{re.escape(em)}\b", text):
                        marker = f"error: {em}"
                        break
                FAILED_LOGS.append((fpath, marker))

            per_level = extract_per_level_fields(text)
            per_level.update(extract_lpm_cycle_class(text))

            key = (
                arch,
                cores,
                meta["trace"],
                meta["l1pf"],
                meta["l2pf"],
                meta["llcpf"],
                meta["l1m"],
                meta["l2m"],
                meta["llcm"],
            )
            metrics = {"ipc": ipc}
            metrics.update(per_level)
            data[key] = metrics

    return data, counts, unparsed_names


def deduce_missing(data):
    """CP8 test-matrix deduction.

    Build the expected matrix = cross-product of observed dimensions, then
    insert SENTINEL_RUNNING placeholder entries for any expected key that
    has no log at all. Keys with a present-but-errored log stay as-is (they
    already carry SENTINEL_ERRORED in their metric values).

    Expected rule (conservative, per-cores-count scoping):
      For each (arch, cores) pair, take the cross-product of:
        traces observed at that (arch, cores)
      × config-tuples (l1pf, l2pf, llcpf, l1m, l2m, llcm) observed at
        that (arch, cores).
      If a resulting key is not in data, insert it with a placeholder
      metrics dict (ipc = -1, no per-level fields). Downstream sheet
      builders already substitute SENTINEL_RUNNING for missing metric keys,
      so inserting just {"ipc": -1} is enough to surface the row.

    Scoping per (arch, cores) rather than globally avoids exploding the
    matrix across unrelated core counts (e.g. 1c runs a different suite
    than 16c). Returns (n_added, n_total_expected).
    """
    from collections import defaultdict as _dd
    per_ac_traces = _dd(set)
    per_ac_configs = _dd(set)

    for key in data.keys():
        arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm = key
        per_ac_traces[(arch, cores)].add(trace)
        per_ac_configs[(arch, cores)].add((l1pf, l2pf, llcpf, l1m, l2m, llcm))

    n_added = 0
    n_expected = 0
    for ac, traces in per_ac_traces.items():
        configs = per_ac_configs[ac]
        for trace in traces:
            for cfg in configs:
                full_key = (ac[0], ac[1], trace) + cfg
                n_expected += 1
                if full_key not in data:
                    data[full_key] = {"ipc": SENTINEL_RUNNING}
                    n_added += 1
    return n_added, n_expected


def validate_ipc_count(outputs_dir, python_ok_count):
    """Cross-check python's IPC extraction against independent grep.

    Returns (grep_count, missing_basenames). Warns if python undercounted.
    """
    try:
        # grep -c per file, then sum; also capture list of files that have the marker
        result = subprocess.run(
            [
                "bash",
                "-c",
                f"grep -rl 'FINAL ROI CORE AVG IPC' {outputs_dir}/core*/ 2>/dev/null || true",
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
        files_with_ipc = [
            ln.strip() for ln in result.stdout.splitlines() if ln.strip()
        ]
    except (subprocess.SubprocessError, OSError):
        return None, []
    grep_count = len(files_with_ipc)
    return grep_count, files_with_ipc


NO_MODEL_TOKEN = "no"


def _is_no_model(l1m, l2m, llcm):
    return l1m == NO_MODEL_TOKEN and l2m == NO_MODEL_TOKEN and llcm == NO_MODEL_TOKEN


def build_sheets_triple(data, field_name):
    """Like build_sheets, but each cell holds a 3-tuple (L1D, L2C, LLC) value list.

    Value for a given (key, trace) is a tuple of 3 scalars (or SENTINEL_RUNNING).
    """
    per_model = defaultdict(lambda: defaultdict(dict))
    per_model_traces = defaultdict(set)
    per_model_pfs = defaultdict(set)

    combined = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    combined_traces = set()
    combined_pfs = defaultdict(set)
    combined_models = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm), metrics in data.items():
        model_combo = f"{l1m}+{l2m}+{llcm}"
        pf_combo = f"L1:{l1pf} L2:{l2pf} LLC:{llcpf}"
        triple = tuple(
            metrics.get((lvl, field_name), SENTINEL_RUNNING) for lvl in LEVELS_TRIPLE
        )

        per_model[model_combo][(arch, cores, pf_combo)][trace] = triple
        per_model_traces[model_combo].add(trace)
        per_model_pfs[(model_combo, arch, cores)].add(pf_combo)

        combined[(arch, cores)][pf_combo][model_combo][trace] = triple
        combined_traces.add(trace)
        combined_pfs[(arch, cores)].add(pf_combo)
        combined_models[(arch, cores, pf_combo)].add(model_combo)

    return {
        "per_model": per_model,
        "per_model_traces": per_model_traces,
        "per_model_pfs": per_model_pfs,
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_pfs": combined_pfs,
        "combined_models": combined_models,
    }


def build_sheets_triple_covbyp(data):
    """Triple-cell builder for 'coverage with bypass crediting'.

    Value = max(load_miss - byp_issued, 0). The downstream coverage formula
    then computes (base - this)/base per-level, giving credit to bypass loads
    as covered misses. If either underlying value missing → SENTINEL_RUNNING.
    """
    per_model = defaultdict(lambda: defaultdict(dict))
    per_model_traces = defaultdict(set)
    per_model_pfs = defaultdict(set)
    combined = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    combined_traces = set()
    combined_pfs = defaultdict(set)
    combined_models = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm), metrics in data.items():
        model_combo = f"{l1m}+{l2m}+{llcm}"
        pf_combo = f"L1:{l1pf} L2:{l2pf} LLC:{llcpf}"
        triple = []
        for lvl in LEVELS_TRIPLE:
            mv = metrics.get((lvl, "load_miss"))
            bv = metrics.get((lvl, "byp_issued"), 0)
            if mv is None:
                triple.append(SENTINEL_RUNNING)
            else:
                triple.append(max(mv - (bv or 0), 0))
        triple = tuple(triple)

        per_model[model_combo][(arch, cores, pf_combo)][trace] = triple
        per_model_traces[model_combo].add(trace)
        per_model_pfs[(model_combo, arch, cores)].add(pf_combo)
        combined[(arch, cores)][pf_combo][model_combo][trace] = triple
        combined_traces.add(trace)
        combined_pfs[(arch, cores)].add(pf_combo)
        combined_models[(arch, cores, pf_combo)].add(model_combo)

    return {
        "per_model": per_model,
        "per_model_traces": per_model_traces,
        "per_model_pfs": per_model_pfs,
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_pfs": combined_pfs,
        "combined_models": combined_models,
    }


def build_sheets(data, metric_key):
    """Like before but value = data[key][metric_key] if present else SENTINEL_RUNNING.

    metric_key is either 'ipc' or a (lvl, field) tuple.
    """
    per_model = defaultdict(lambda: defaultdict(dict))
    per_model_traces = defaultdict(set)
    per_model_pfs = defaultdict(set)

    combined = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    combined_traces = set()
    combined_pfs = defaultdict(set)
    combined_models = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm), metrics in data.items():
        model_combo = f"{l1m}+{l2m}+{llcm}"
        pf_combo = f"L1:{l1pf} L2:{l2pf} LLC:{llcpf}"
        val = metrics.get(metric_key, SENTINEL_RUNNING)

        per_model[model_combo][(arch, cores, pf_combo)][trace] = val
        per_model_traces[model_combo].add(trace)
        per_model_pfs[(model_combo, arch, cores)].add(pf_combo)

        combined[(arch, cores)][pf_combo][model_combo][trace] = val
        combined_traces.add(trace)
        combined_pfs[(arch, cores)].add(pf_combo)
        combined_models[(arch, cores, pf_combo)].add(model_combo)

    return {
        "per_model": per_model,
        "per_model_traces": per_model_traces,
        "per_model_pfs": per_model_pfs,
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_pfs": combined_pfs,
        "combined_models": combined_models,
    }


def base_sort_key(pf_combo):
    """Sort key so that the 'base' pf combo lands LAST in each group.

    CP1 base = all pf = 'no' (L1:no L2:no LLC:no). Fallback: lexicographic
    last by prefetcher string (highest ASCII sorts last in plain sort).
    """
    is_all_no = pf_combo == "L1:no L2:no LLC:no"
    # Tuple: (is_all_no, pf_combo). is_all_no True -> 1 which is > 0, so
    # the all-no row naturally sorts after any False (0) entry under default
    # ascending sort. Tiebreak on name.
    return (1 if is_all_no else 0, pf_combo)


HEADER_COLS_PER_MODEL = ["Arch", "Cores", "Pf \\ Traces"]
HEADER_COLS_COMBINED = ["Arch", "Cores", "Pf", "Model \\ Traces"]


def _col_letter(col_idx):
    s = ""
    n = col_idx
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def _write_header(ws, row, labels, traces):
    for i, lab in enumerate(labels):
        ws.cell(row=row, column=1 + i, value=lab)
    start = len(labels) + 1
    for i, trace in enumerate(traces):
        ws.cell(row=row, column=start + i, value=trace)


def _set_ipc_cell(ws, row, col, val):
    cell = ws.cell(row=row, column=col, value=val)
    if isinstance(val, float) and val >= 0:
        cell.number_format = "0.0000"
    return cell


def _speedup_formula(src_col_letter, src_row, base_row):
    return (
        f'=IF(AND({src_col_letter}{src_row}>=0,{src_col_letter}{base_row}>0),'
        f'{src_col_letter}{src_row}/{src_col_letter}{base_row},"")'
    )


def _coverage_formula(src_col_letter, src_row, base_row):
    # (base - src) / base — fraction of misses avoided vs no-model base
    return (
        f'=IF(AND({src_col_letter}{src_row}>=0,{src_col_letter}{base_row}>0),'
        f'({src_col_letter}{base_row}-{src_col_letter}{src_row})'
        f'/{src_col_letter}{base_row},"")'
    )


def _iter_groups(sh):
    """Yield ((arch, cores, pf_combo), ordered_models) for every data group.

    Models within a (arch,cores,pf) group are sorted with 'no+no+no' LAST
    so the base config lands at the end and the speedup formula gets a
    fixed relative reference.
    """
    arch_cores_list = sorted(sh["combined"].keys())
    for (arch, cores) in arch_cores_list:
        pfs = sorted(sh["combined_pfs"][(arch, cores)], key=base_sort_key)
        for pf_combo in pfs:
            models = sorted(sh["combined_models"][(arch, cores, pf_combo)])
            non_base = [m for m in models if m != "no+no+no"]
            base_models = [m for m in models if m == "no+no+no"]
            ordered = non_base + base_models
            yield (arch, cores, pf_combo), ordered


def _write_single_block(
    ws,
    sh,
    start_row,
    formula_kind,
    section_title,
):
    """Single-cell-per-trace block (used by IPC). Returns next_row.

    Layout: title row → for each (arch,cores,pf) group: header row,
    data rows (models, base last), blank row. 3-row gap. Mirror block
    with repeated headers per group and the formula.
    """
    fmt_fn = _speedup_formula if formula_kind == "speedup" else _coverage_formula
    mirror_word = "SPEEDUP" if formula_kind == "speedup" else "COVERAGE"
    traces = sorted(sh["combined_traces"])
    n_traces = len(traces)
    n_label_cols = len(HEADER_COLS_COMBINED)
    first_trace_col = n_label_cols + 1

    row = start_row
    ws.cell(row=row, column=1, value=section_title)
    row += 1

    main_row_of = {}
    base_row_of = {}

    for (arch, cores, pf_combo), ordered_models in _iter_groups(sh):
        _write_header(ws, row, HEADER_COLS_COMBINED, traces)
        row += 1
        for model_combo in ordered_models:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=pf_combo)
            ws.cell(row=row, column=4, value=model_combo)
            for i, trace in enumerate(traces):
                val = sh["combined"][(arch, cores)][pf_combo][model_combo].get(
                    trace, SENTINEL_RUNNING
                )
                _set_ipc_cell(ws, row, first_trace_col + i, val)
            main_row_of[(arch, cores, pf_combo, model_combo)] = row
            if model_combo == "no+no+no":
                base_row_of[(arch, cores, pf_combo)] = row
            row += 1
        row += 1  # gap between groups

    # 3-row gap before speedup block
    row += 2
    ws.cell(row=row, column=1, value=f"{mirror_word} — {section_title}")
    row += 1

    for (arch, cores, pf_combo), ordered_models in _iter_groups(sh):
        _write_header(ws, row, HEADER_COLS_COMBINED, traces)
        row += 1
        base_r = base_row_of.get((arch, cores, pf_combo))
        for model_combo in ordered_models:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=pf_combo)
            ws.cell(row=row, column=4, value=model_combo)
            src_row = main_row_of[(arch, cores, pf_combo, model_combo)]
            for i in range(n_traces):
                col = get_column_letter(first_trace_col + i)
                if base_r is None:
                    ws.cell(row=row, column=first_trace_col + i, value="")
                else:
                    formula = fmt_fn(col, src_row, base_r)
                    cell = ws.cell(row=row, column=first_trace_col + i, value=formula)
                    cell.number_format = "0.0000"
            row += 1
        row += 1

    return row


def _write_metric_sheets(wb, sh, metric_label, formula_kind="speedup"):
    """Emit the single combined IPC-style sheet (no per-model sheets)."""
    combined_title = f"{metric_label} ALL (comb)"[:31]
    ws = wb.create_sheet(title=combined_title)
    _write_single_block(
        ws,
        sh,
        start_row=1,
        formula_kind=formula_kind,
        section_title=f"{metric_label}: all models combined (base = no-model per PF)",
    )


def _write_triple_group_header(ws, row, traces, first_trace_col):
    """Two-row header: row N has label cols + merged trace names,
    row N+1 has L1D/L2C/LLC sub-labels under each trace."""
    for i, lab in enumerate(HEADER_COLS_COMBINED):
        ws.cell(row=row, column=1 + i, value=lab)
    for ti, trace in enumerate(traces):
        c0 = first_trace_col + ti * 3
        cell = ws.cell(row=row, column=c0, value=trace)
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 2)
    for ti in range(len(traces)):
        c0 = first_trace_col + ti * 3
        for li, lvl in enumerate(LEVELS_TRIPLE):
            ws.cell(row=row + 1, column=c0 + li, value=lvl)


def _write_triple_block(ws, sh, start_row, formula_kind, section_title):
    """Triple-cell (L1D/L2C/LLC) block. Returns next_row.

    Headers are repeated at every (arch, cores, pf) group, in both the
    main region and the speedup/coverage mirror region.
    """
    fmt_fn = _speedup_formula if formula_kind == "speedup" else _coverage_formula
    mirror_word = "SPEEDUP" if formula_kind == "speedup" else "COVERAGE"
    traces = sorted(sh["combined_traces"])
    n_traces = len(traces)
    n_label_cols = len(HEADER_COLS_COMBINED)
    first_trace_col = n_label_cols + 1

    row = start_row
    ws.cell(row=row, column=1, value=section_title)
    row += 1

    main_row_of = {}
    base_row_of = {}

    for (arch, cores, pf_combo), ordered_models in _iter_groups(sh):
        _write_triple_group_header(ws, row, traces, first_trace_col)
        row += 2  # header + L1/L2/LLC sub-row
        for model_combo in ordered_models:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=pf_combo)
            ws.cell(row=row, column=4, value=model_combo)
            for ti, trace in enumerate(traces):
                triple = sh["combined"][(arch, cores)][pf_combo][model_combo].get(
                    trace, (SENTINEL_RUNNING,) * 3
                )
                c0 = first_trace_col + ti * 3
                for li in range(3):
                    v = triple[li]
                    cell = ws.cell(row=row, column=c0 + li, value=v)
                    if isinstance(v, float) and v >= 0:
                        cell.number_format = "0.0000"
            main_row_of[(arch, cores, pf_combo, model_combo)] = row
            if model_combo == "no+no+no":
                base_row_of[(arch, cores, pf_combo)] = row
            row += 1
        row += 1  # gap between groups

    # 3-row gap + mirror region title
    row += 2
    ws.cell(row=row, column=1, value=f"{mirror_word} — {section_title}")
    row += 1

    for (arch, cores, pf_combo), ordered_models in _iter_groups(sh):
        _write_triple_group_header(ws, row, traces, first_trace_col)
        row += 2
        base_r = base_row_of.get((arch, cores, pf_combo))
        for model_combo in ordered_models:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=pf_combo)
            ws.cell(row=row, column=4, value=model_combo)
            src_row = main_row_of[(arch, cores, pf_combo, model_combo)]
            for ti in range(n_traces):
                c0 = first_trace_col + ti * 3
                for li in range(3):
                    col = get_column_letter(c0 + li)
                    if base_r is None:
                        ws.cell(row=row, column=c0 + li, value="")
                    else:
                        formula = fmt_fn(col, src_row, base_r)
                        cell = ws.cell(row=row, column=c0 + li, value=formula)
                        cell.number_format = "0.0000"
            row += 1
        row += 1

    return row


def _write_triple_family_sheet(wb, sh, metric_label, formula_kind="speedup"):
    """Emit a single-block triple-cell sheet for a metric family."""
    title = f"{metric_label} ALL (comb)"[:31]
    ws = wb.create_sheet(title=title)
    _write_triple_block(
        ws,
        sh,
        start_row=1,
        formula_kind=formula_kind,
        section_title=f"{metric_label}: per PF×model, L1D/L2C/LLC triple",
    )


def build_sheets_multilevel(data, field_per_sub):
    """Variant of build_sheets_triple where each cell holds a tuple of
    (len(LEVELS_TRIPLE) * len(field_per_sub)) values.

    field_per_sub: ordered list of field names to pull per level.
    Order of packed tuple: for lvl in LEVELS_TRIPLE: for f in field_per_sub: val.
    """
    per_model = defaultdict(lambda: defaultdict(dict))
    per_model_traces = defaultdict(set)
    per_model_pfs = defaultdict(set)
    combined = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    combined_traces = set()
    combined_pfs = defaultdict(set)
    combined_models = defaultdict(set)

    for (arch, cores, trace, l1pf, l2pf, llcpf, l1m, l2m, llcm), metrics in data.items():
        model_combo = f"{l1m}+{l2m}+{llcm}"
        pf_combo = f"L1:{l1pf} L2:{l2pf} LLC:{llcpf}"
        packed = []
        for lvl in LEVELS_TRIPLE:
            for f in field_per_sub:
                packed.append(metrics.get((lvl, f), SENTINEL_RUNNING))
        packed = tuple(packed)
        per_model[model_combo][(arch, cores, pf_combo)][trace] = packed
        per_model_traces[model_combo].add(trace)
        per_model_pfs[(model_combo, arch, cores)].add(pf_combo)
        combined[(arch, cores)][pf_combo][model_combo][trace] = packed
        combined_traces.add(trace)
        combined_pfs[(arch, cores)].add(pf_combo)
        combined_models[(arch, cores, pf_combo)].add(model_combo)

    return {
        "per_model": per_model,
        "per_model_traces": per_model_traces,
        "per_model_pfs": per_model_pfs,
        "combined": combined,
        "combined_traces": combined_traces,
        "combined_pfs": combined_pfs,
        "combined_models": combined_models,
    }


def _write_multilevel_header(ws, row, traces, first_trace_col, sub_labels):
    """Three-row header: merged trace names, merged level names under each
    trace (spanning len(sub_labels) cells each), and sub-labels under the
    levels.
    """
    width = 3 * len(sub_labels)
    for i, lab in enumerate(HEADER_COLS_COMBINED):
        ws.cell(row=row, column=1 + i, value=lab)
    for ti, trace in enumerate(traces):
        c0 = first_trace_col + ti * width
        cell = ws.cell(row=row, column=c0, value=trace)
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=row, start_column=c0, end_row=row, end_column=c0 + width - 1
        )
    # row+1: level labels (L1D / L2C / LLC), each spanning len(sub_labels) cols
    for ti in range(len(traces)):
        c0 = first_trace_col + ti * width
        for li, lvl in enumerate(LEVELS_TRIPLE):
            sub_c0 = c0 + li * len(sub_labels)
            cell = ws.cell(row=row + 1, column=sub_c0, value=lvl)
            cell.alignment = Alignment(horizontal="center")
            if len(sub_labels) > 1:
                ws.merge_cells(
                    start_row=row + 1,
                    start_column=sub_c0,
                    end_row=row + 1,
                    end_column=sub_c0 + len(sub_labels) - 1,
                )
    # row+2: sub-labels repeated under each level
    for ti in range(len(traces)):
        c0 = first_trace_col + ti * width
        for li in range(3):
            for si, sub_lab in enumerate(sub_labels):
                ws.cell(
                    row=row + 2,
                    column=c0 + li * len(sub_labels) + si,
                    value=sub_lab,
                )


def _write_coverage_multicol_block(ws, data, start_row, field_per_sub, sub_labels,
                                   section_title, percent_format=True):
    """Emit a coverage block where each trace header spans 3*len(sub_labels) cols.

    Main region: raw counts. Mirror region: coverage formula (base-this)/base
    applied ONLY to sub-columns whose label equals 'Misses'; the rest are blank.
    Percent format on mirror region cells when percent_format=True.
    """
    sh = build_sheets_multilevel(data, field_per_sub)
    traces = sorted(sh["combined_traces"])
    n_sub = len(sub_labels)
    width_per_trace = 3 * n_sub
    n_label_cols = len(HEADER_COLS_COMBINED)
    first_trace_col = n_label_cols + 1

    row = start_row
    ws.cell(row=row, column=1, value=section_title)
    row += 1

    main_row_of = {}
    base_row_of = {}

    for (arch, cores, pf_combo), ordered_models in _iter_groups(sh):
        _write_multilevel_header(ws, row, traces, first_trace_col, sub_labels)
        row += 3  # 3-row header (trace / level / sublabel)
        for model_combo in ordered_models:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=pf_combo)
            ws.cell(row=row, column=4, value=model_combo)
            for ti, trace in enumerate(traces):
                packed = sh["combined"][(arch, cores)][pf_combo][model_combo].get(
                    trace, (SENTINEL_RUNNING,) * (3 * n_sub)
                )
                c0 = first_trace_col + ti * width_per_trace
                for idx in range(3 * n_sub):
                    v = packed[idx]
                    cell = ws.cell(row=row, column=c0 + idx, value=v)
                    if isinstance(v, (int, float)) and v >= 0:
                        cell.number_format = "0"
            main_row_of[(arch, cores, pf_combo, model_combo)] = row
            if model_combo == "no+no+no":
                base_row_of[(arch, cores, pf_combo)] = row
            row += 1
        row += 1  # gap between groups

    # 3-row gap + mirror region title
    row += 2
    ws.cell(row=row, column=1, value=f"COVERAGE — {section_title}  (=(base-eval)/base)")
    row += 1

    # Indexes by role. Miss → (base-eval)/base.  ByP → eval/base (fraction of
    # base-misses absorbed by bypass). Hits → blank (no semantic meaning).
    miss_sub_idxs = [i for i, s in enumerate(sub_labels) if s.lower().startswith("miss")]
    byp_sub_idxs = [i for i, s in enumerate(sub_labels) if s.lower() == "byp"]
    # Locate the corresponding miss sub-idx for each level so ByP formulas
    # can reference the base row's MISS column (not the base row's ByP column).
    miss_relative_idx = miss_sub_idxs[0] if miss_sub_idxs else None

    for (arch, cores, pf_combo), ordered_models in _iter_groups(sh):
        _write_multilevel_header(ws, row, traces, first_trace_col, sub_labels)
        row += 3
        base_r = base_row_of.get((arch, cores, pf_combo))
        for model_combo in ordered_models:
            ws.cell(row=row, column=1, value=arch)
            ws.cell(row=row, column=2, value=cores)
            ws.cell(row=row, column=3, value=pf_combo)
            ws.cell(row=row, column=4, value=model_combo)
            src_row = main_row_of[(arch, cores, pf_combo, model_combo)]
            for ti in range(len(traces)):
                c0 = first_trace_col + ti * width_per_trace
                for li in range(3):
                    for si in range(n_sub):
                        idx = c0 + li * n_sub + si
                        col = get_column_letter(idx)
                        if si in miss_sub_idxs and base_r is not None:
                            # Coverage = (base_miss - eval_miss) / base_miss
                            formula = _coverage_formula(col, src_row, base_r)
                            cell = ws.cell(row=row, column=idx, value=formula)
                            cell.number_format = "0.00%" if percent_format else "0.0000"
                        elif si in byp_sub_idxs and base_r is not None and miss_relative_idx is not None:
                            # ByP cell = (base_miss - eval_byp) / base_miss.
                            # Base misses include the eval's ByP count because
                            # ChampSim increments total on byp AND on miss, so
                            # the standard miss formula captures hit-coverage
                            # and the ByP formula captures bypass-coverage —
                            # the two sum to total miss coverage. Same shape
                            # as the Miss cell; only the eval column changes.
                            byp_col = get_column_letter(idx)  # eval ByP col
                            base_miss_idx = c0 + li * n_sub + miss_relative_idx
                            base_miss_col = get_column_letter(base_miss_idx)
                            formula = (
                                f'=IF(AND({byp_col}{src_row}>=0,'
                                f'{base_miss_col}{base_r}>0),'
                                f'({base_miss_col}{base_r}-{byp_col}{src_row})'
                                f'/{base_miss_col}{base_r},"")'
                            )
                            cell = ws.cell(row=row, column=idx, value=formula)
                            cell.number_format = "0.00%" if percent_format else "0.0000"
                        else:
                            ws.cell(row=row, column=idx, value="")
            row += 1
        row += 1

    return row


def _write_consolidated_coverage_sheet(wb, data):
    """Single Coverage sheet with two sibling blocks:

    1) ByP-aware: per trace header → 9 cols = L1D/L2C/LLC × Hits/ByP/Misses,
       sourced from load_wByP_hit / load_wByP_byp / load_wByP_miss.
       Coverage formula ((base-eval)/base) on the Misses sub-cols.

    2) STD: per trace header → 6 cols = L1D/L2C/LLC × Hits/Misses,
       sourced from load_hit / load_miss. Coverage formula on Misses.

    Coverage displayed as percentage (0.00%).
    """
    ws = wb.create_sheet(title="Coverage")
    row = 1
    ws.cell(
        row=row,
        column=1,
        value="COVERAGE — ByP-aware (9-col: Hits/ByP/Misses per level), then STD (6-col: Hits/Misses per level)",
    )
    row += 2

    # Block 1: ByP-aware
    row = _write_coverage_multicol_block(
        ws,
        data,
        row,
        field_per_sub=["load_wByP_hit", "load_wByP_byp", "load_wByP_miss"],
        sub_labels=["Hits", "ByP", "Misses"],
        section_title="ByP-aware (load_wByP_hit / load_wByP_byp / load_wByP_miss)",
    )
    row += 3

    # Block 2: STD
    ws.cell(row=row, column=1, value="STANDARD (no ByP — reported load_hit / load_miss)")
    row += 2
    row = _write_coverage_multicol_block(
        ws,
        data,
        row,
        field_per_sub=["load_hit", "load_miss"],
        sub_labels=["Hits", "Misses"],
        section_title="STD (load_hit / load_miss)",
    )


def write_xlsx(data, input_basename):
    wb = Workbook()
    wb.remove(wb.active)

    # Order: IPC → MPKI → Coverage (consolidated) → APC → kappa → phi
    #        → C-AMAT → MST → LPMR_std → LPMR_byp → Loads/Miss/MSHR/Byp
    # Per user 2026-04-14: drop per-model sheets, drop CPA (never logged).

    # 1) IPC (single-cell)
    sh_ipc = build_sheets(data, "ipc")
    _write_metric_sheets(wb, sh_ipc, "IPC")

    # 2) MPKI (triple-cell)
    sh_mpki = build_sheets_triple(data, "MPKI")
    _write_triple_family_sheet(wb, sh_mpki, "MPKI", formula_kind="speedup")

    # 3) Consolidated Coverage sheet (ByP-credited stacked, then STD)
    _write_consolidated_coverage_sheet(wb, data)

    # 4) LPM key metrics — order: APC, kappa, phi, C-AMAT, MST, LPMR_std, LPMR_byp.
    #    CPA excluded: tracked by ChampSim but never emitted to logs.
    for col in ("APC", "kappa", "phi", "C-AMAT", "MST", "LPMR_std", "LPMR_byp"):
        sh_lpm = build_sheets_triple(data, "LPM_" + col)
        _write_triple_family_sheet(wb, sh_lpm, f"LPM {col}", formula_kind="speedup")

    # 5) Remaining metric families (Loads, Miss, MSHR caps, Byp counters)
    for sheet_label, field in CP5_FAMILIES:
        if sheet_label == "MPKI":
            continue  # already emitted above
        sh_tri = build_sheets_triple(data, field)
        _write_triple_family_sheet(wb, sh_tri, sheet_label, formula_kind="speedup")

    ts = int(time.time())
    out_name = f"{input_basename}_CP7b_{ts}.xlsx"
    out_path = os.path.join(OUTPUT_DIR, out_name)
    wb.save(out_path)
    return out_path


def _debug_dump(data, section):
    """Pretty-print a slice of the hashmap for --debug <section>."""
    section = section.lower()
    print(f"=== DEBUG: {section} ===")
    shown = 0
    for key, metrics in sorted(data.items()):
        if section == "ipc":
            val = metrics.get("ipc", SENTINEL_RUNNING)
            print(f"  {key} -> ipc={val}")
        elif section in ("loads", "miss", "mpki", "byp"):
            field_map = {
                "loads": "loads",
                "miss": "load_miss",
                "mpki": "MPKI",
                "byp": "byp_req",
            }
            fld = field_map[section]
            triple = tuple(metrics.get((lvl, fld), SENTINEL_RUNNING) for lvl in LEVELS_TRIPLE)
            print(f"  {key} -> {fld}{triple}")
        elif section == "cycle_class":
            lpm_keys = [k for k in metrics if isinstance(k, tuple) and k[1] in
                        ("APC", "kappa", "phi", "C-AMAT", "MST", "LPMR_std", "LPMR_byp")]
            print(f"  {key} -> {len(lpm_keys)} LPM fields")
        elif section == "raw":
            print(f"  {key} -> {metrics}")
        else:
            print(f"  (unknown section '{section}', try: ipc|loads|miss|mpki|byp|cycle_class|raw)")
            return
        shown += 1
        if shown >= 50:
            print(f"  ... ({len(data) - 50} more entries suppressed)")
            break


def main():
    args = sys.argv[1:]
    debug_section = None
    if "--debug" in args:
        i = args.index("--debug")
        if i + 1 >= len(args):
            print("usage: aggregator.py <outputs-dir> [--debug <section>]", file=sys.stderr)
            sys.exit(2)
        debug_section = args[i + 1]
        args = args[:i] + args[i + 2:]
    if len(args) != 1:
        print("usage: aggregator.py <outputs-dir> [--debug <section>]", file=sys.stderr)
        sys.exit(2)
    outputs_dir = args[0]
    if not os.path.isdir(outputs_dir):
        print(f"not a dir: {outputs_dir}", file=sys.stderr)
        sys.exit(2)

    input_basename = os.path.basename(os.path.abspath(outputs_dir.rstrip("/")))

    data, counts, unparsed_names = collect(outputs_dir)
    n_added, n_expected = deduce_missing(data)
    counts["deduced_missing"] = n_added
    counts["expected_matrix"] = n_expected

    if debug_section is not None:
        _debug_dump(data, debug_section)

    out_path = write_xlsx(data, input_basename)

    print(
        f"logs={counts['logs']} parsed={counts['parsed']} "
        f"ok={counts['ok']} running={counts['running']} errored={counts['errored']} "
        f"deduced_missing={counts['deduced_missing']}/{counts['expected_matrix']} "
        f"out={os.path.basename(out_path)}"
    )

    grep_count, files_with_ipc = validate_ipc_count(outputs_dir, counts["ok"])
    if grep_count is not None:
        if grep_count != counts["ok"]:
            print(
                f"WARN: grep found FINAL ROI in {grep_count} files but python "
                f"extracted {counts['ok']}  (delta={grep_count - counts['ok']})"
            )
            py_ok_basenames = set()
            # re-walk and record which logs python classified as ok
            for entry in os.listdir(outputs_dir):
                core_match = CORE_DIR_RE.match(entry)
                if not core_match:
                    continue
                cd = os.path.join(outputs_dir, entry)
                if not os.path.isdir(cd):
                    continue
                for fn in os.listdir(cd):
                    if not fn.endswith(".log"):
                        continue
                    meta = parse_log_filename(fn)
                    if meta is None:
                        continue
                    text = read_log(os.path.join(cd, fn))
                    ipc, sent = extract_ipc(text)
                    if sent == 0:
                        py_ok_basenames.add(os.path.abspath(os.path.join(cd, fn)))
            grep_set = {os.path.abspath(p) for p in files_with_ipc}
            missing = sorted(grep_set - py_ok_basenames)
            for p in missing[:20]:
                print(f"  MISSING: {os.path.basename(p)}")
            if len(missing) > 20:
                print(f"  ... and {len(missing) - 20} more")
        else:
            print(f"OK: grep/python IPC count match ({grep_count})")

    if unparsed_names:
        print(f"WARN: {len(unparsed_names)} log filenames did not match parser")
        for n in unparsed_names[:10]:
            print(f"  UNPARSED: {os.path.basename(n)}")
        if len(unparsed_names) > 10:
            print(f"  ... and {len(unparsed_names) - 10} more")


if __name__ == "__main__":
    main()
