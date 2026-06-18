#!/usr/bin/env python3
"""Excel-native paper figures (openpyxl bar charts) for aggregator_parallel.

LIVE-FORMULA design (per user): figures are NOT computed in Python. Each figure
sheet writes a clean staging table whose cells are Excel formulas referencing the
already-built 'IPC' sheet's SPEEDUP-mirror GM column (geomean over traces). The
chart then references that staging table. Edit IPC -> table + chart auto-update.

Speedup methodology (already encoded in the IPC speedup mirror, confirmed by user):
  speedup(repl,PF,technique,trace) = IPC(repl,PF,technique,trace)
                                     / IPC(repl, no-Pref, Base-technique, trace)
  => PER-REPL OWN BASE; no-Pref+Base = 1.000. GM column = geomean over traces.

Figures built here:
  F1 (per-PF, LRU): X=cores, one series per prefetcher, value = MARS speedup GM.
      Establishing figure ("show once") justifying averaging the prefetchers.
  F2 (LRU technique): summary table per core with a no-Pref "Base" row and an
      "Avg-PF" row, columns Base/HERMES/HMP/TTP/MARS. Chart on the Avg-PF row.
"""

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.utils import get_column_letter

# ── IPC-sheet vocabulary ──
IPC_SHEET = "IPC"
NO_PF_COMBO = "L1:no L2:no L3:no"
TECH_OCP = {"Base", "HERMES", "HMP", "TTP"}   # MARS = the remaining technique label

# Prefetcher combos in figure (light->dark) order + display labels.
PF_FIG_ORDER = [
    "L1:no L2:spp L3:no",
    "L1:no L2:Zion1 L3:no",
    "L1:no L2:bingo_dpc3 L3:no",
    "L1:no L2:berti_c L3:no",
]
PF_COMBO_LABEL = {
    NO_PF_COMBO: "no-Pref",
    "L1:no L2:spp L3:no": "SPP",
    "L1:no L2:Zion1 L3:no": "Zion",
    "L1:no L2:bingo_dpc3 L3:no": "Bingo",
    "L1:no L2:berti_c L3:no": "Berti",
}

CORE_ORDER = [1, 4, 8, 16]
CORE_LABEL = {1: "1 Core", 2: "2 Core", 4: "4 Core", 8: "8 Core", 16: "16 Core"}


# ── Style (paper) ──
# white(bordered) -> light gray -> diagonal hatch -> dark gray -> black.
# last series forced solid black ("mine" = MARS, rightmost).
_PALETTE = [
    ("solid", "FFFFFF"),
    ("solid", "D9D9D9"),
    ("hatch", "ltUpDiag"),
    ("solid", "A6A6A6"),
    ("solid", "595959"),
    ("solid", "000000"),
]


def _fill_scheme(n):
    if n <= 1:
        return [("solid", "000000")]
    light = _PALETTE[:-1]
    chosen = light[:max(n - 1, 0)]
    while len(chosen) < n - 1:
        chosen.append(("solid", "808080"))
    return chosen + [("solid", "000000")]


def _gray_gradient(n):
    return _fill_scheme(n)


def _style_series(chart, scheme, border="000000", border_w=9525):
    for idx, ser in enumerate(chart.series):
        kind = scheme[idx % len(scheme)]
        gp = GraphicalProperties()
        if kind[0] == "hatch":
            gp.pattFill = PatternFillProperties(
                prst=kind[1],
                fgClr=ColorChoice(srgbClr="000000"),
                bgClr=ColorChoice(srgbClr="FFFFFF"))
        else:
            gp.solidFill = kind[1]
        gp.line = LineProperties(solidFill=border, w=border_w)
        ser.graphicalProperties = gp


def _grouped_bar(ws, y_title, cat_ref, data_ref, anchor,
                 y_min=1.0, y_max=None, y_fmt="0.00",
                 colors=None, width_cm=17.8, height_cm=4.3, gap=60):
    """Clustered grouped-bar chart, paper style: no chart title, Y-axis title,
    horizontal gridlines kept, decimal axis."""
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "clustered"
    ch.overlap = 0
    ch.gapWidth = gap
    ch.title = None
    ch.height = height_cm
    ch.width = width_cm
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cat_ref)
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.y_axis.numFmt = y_fmt
    ch.y_axis.scaling.min = y_min
    if y_max is not None:
        ch.y_axis.scaling.max = y_max
    ch.y_axis.title = y_title
    ch.legend.position = "t"
    if colors:
        _style_series(ch, colors)
    ws.add_chart(ch, anchor)
    return ch


# ════════════════════════════════════════════════════════════════════════
# Scan the built IPC sheet's SPEEDUP-mirror block for GM-cell addresses
# ════════════════════════════════════════════════════════════════════════

def _scan_ipc_speedup(ws):
    """Locate the SPEEDUP-mirror GM block in the IPC sheet.

    Returns dict:
      tech_cols : {technique_label -> GM column index}
      techs     : ordered technique labels (Base,HERMES,HMP,TTP,MARS)
      mars      : the MARS/forwarding technique label (non-OCP)
      row_of    : {(cores:int, repl:str, pf_combo:str) -> speedup data row}
    or None if the block can't be found.
    """
    sp_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith("SPEEDUP"):
            sp_row = r
            break
    if sp_row is None:
        return None

    trace_row = sp_row + 1      # merged trace names + GM/AVG headers
    hdr2 = sp_row + 2           # sub-col labels (technique names) + Cores/Repl/PF

    gm_col = avg_col = None
    for c in range(4, ws.max_column + 1):
        v = ws.cell(row=trace_row, column=c).value
        if v == "GM" and gm_col is None:
            gm_col = c
        elif v == "AVG" and gm_col is not None and avg_col is None:
            avg_col = c
            break
    if gm_col is None:
        return None
    n_sub = (avg_col - gm_col) if avg_col else 5

    techs = []
    tech_cols = {}
    for i in range(n_sub):
        lbl = ws.cell(row=hdr2, column=gm_col + i).value
        if lbl in (None, ""):
            break
        techs.append(lbl)
        tech_cols[lbl] = gm_col + i
    mars = next((t for t in techs if t not in TECH_OCP), techs[-1] if techs else None)

    row_of = {}
    for r in range(hdr2 + 1, ws.max_row + 1):
        cs = ws.cell(row=r, column=1).value
        rp = ws.cell(row=r, column=2).value
        pf = ws.cell(row=r, column=3).value
        if cs is not None and rp is not None and pf is not None:
            try:
                row_of[(int(cs), str(rp), str(pf))] = r
            except (TypeError, ValueError):
                pass

    return {"tech_cols": tech_cols, "techs": techs, "mars": mars, "row_of": row_of}


def _gm_ref(scan, cores, repl, pf_combo, tech):
    """Excel address string 'IPC!<col><row>' of the speedup GM cell, or None."""
    r = scan["row_of"].get((cores, repl, pf_combo))
    c = scan["tech_cols"].get(tech)
    if r is None or c is None:
        return None
    return f"{IPC_SHEET}!{get_column_letter(c)}{r}"


# ════════════════════════════════════════════════════════════════════════
# F1 — per-PF LRU MARS speedup, X = cores (live formulas)
# ════════════════════════════════════════════════════════════════════════

def build_F1(wb, scan, cores_present):
    repl = "lru"
    cores = [c for c in CORE_ORDER if c in cores_present]
    mars = scan["mars"]
    pfs = [p for p in PF_FIG_ORDER
           if any(_gm_ref(scan, c, repl, p, mars) for c in cores)]
    if not cores or not pfs or mars is None:
        return None

    ws = wb.create_sheet("FIG_F1_perPF")
    ws["A1"] = "F1: per-PF LRU MARS speedup vs no-pref base (live =IPC! GM refs)"
    hdr = 3
    ws.cell(row=hdr, column=1, value="Cores")
    for j, p in enumerate(pfs):
        ws.cell(row=hdr, column=2 + j, value=PF_COMBO_LABEL.get(p, p))
    for i, c in enumerate(cores):
        r = hdr + 1 + i
        ws.cell(row=r, column=1, value=CORE_LABEL.get(c, f"{c} Core"))
        for j, p in enumerate(pfs):
            ref = _gm_ref(scan, c, repl, p, mars)
            ws.cell(row=r, column=2 + j, value=(f"={ref}" if ref else None))

    n_rows = len(cores)
    n_ser = len(pfs)
    cat = Reference(ws, min_col=1, min_row=hdr + 1, max_row=hdr + n_rows)
    dat = Reference(ws, min_col=2, max_col=1 + n_ser, min_row=hdr, max_row=hdr + n_rows)
    _grouped_bar(ws, "Normalized Speedup", cat, dat, "H3",
                 y_min=1.0, colors=_fill_scheme(n_ser))
    return ws


# ════════════════════════════════════════════════════════════════════════
# F2 — LRU technique comparison, X = cores (Avg-over-PF speedup, live formulas)
# Summary table: per core a no-Pref "Base" row + an "Avg-PF" row, cols=techniques.
# Chart plots the Avg-PF row across cores, one series per technique.
# ════════════════════════════════════════════════════════════════════════

def build_F2(wb, scan, cores_present):
    repl = "lru"
    cores = [c for c in CORE_ORDER if c in cores_present]
    techs = scan["techs"]
    if not cores or not techs:
        return None

    ws = wb.create_sheet("FIG_F2_LRU_tech")
    ws["A1"] = "F2: LRU technique speedup (per-repl own base, live =IPC! GM refs)"

    hdr = 3
    ws.cell(row=hdr, column=1, value="Cores")
    ws.cell(row=hdr, column=2, value="Row")
    for j, t in enumerate(techs):
        ws.cell(row=hdr, column=3 + j, value=t)

    # Two rows per core: no-Pref "Base" speedups, and "Avg-PF" speedups.
    avg_rows = []   # rows holding the Avg-PF series (charted)
    r = hdr + 1
    for c in cores:
        # no-Pref Base row
        ws.cell(row=r, column=1, value=CORE_LABEL.get(c, f"{c} Core"))
        ws.cell(row=r, column=2, value="no-Pref")
        for j, t in enumerate(techs):
            ref = _gm_ref(scan, c, repl, NO_PF_COMBO, t)
            ws.cell(row=r, column=3 + j, value=(f"={ref}" if ref else None))
        r += 1
        # Avg-PF row (mean of the 4 prefetchers' speedup for each technique)
        ws.cell(row=r, column=1, value=CORE_LABEL.get(c, f"{c} Core"))
        ws.cell(row=r, column=2, value="Avg-PF")
        for j, t in enumerate(techs):
            refs = [_gm_ref(scan, c, repl, p, t) for p in PF_FIG_ORDER]
            refs = [x for x in refs if x]
            ws.cell(row=r, column=3 + j,
                    value=(f"=AVERAGE({','.join(refs)})" if refs else None))
        avg_rows.append(r)
        r += 1

    # Chart: X = cores (Avg-PF rows), one series per technique.
    # Build a compact contiguous chart-feed block so categories are the cores.
    feed_hdr = r + 2
    ws.cell(row=feed_hdr, column=1, value="(chart feed: Avg-PF speedup per core)")
    fr = feed_hdr + 1
    ws.cell(row=fr, column=1, value="Cores")
    for j, t in enumerate(techs):
        ws.cell(row=fr, column=2 + j, value=t)
    for i, c in enumerate(cores):
        rr = fr + 1 + i
        ws.cell(row=rr, column=1, value=CORE_LABEL.get(c, f"{c} Core"))
        for j in range(len(techs)):
            src = ws.cell(row=avg_rows[i], column=3 + j).coordinate
            ws.cell(row=rr, column=2 + j, value=f"={src}")

    n_rows = len(cores)
    n_ser = len(techs)
    cat = Reference(ws, min_col=1, min_row=fr + 1, max_row=fr + n_rows)
    dat = Reference(ws, min_col=2, max_col=1 + n_ser, min_row=fr, max_row=fr + n_rows)
    _grouped_bar(ws, "Normalized Speedup", cat, dat, "H3",
                 y_min=1.0, colors=_fill_scheme(n_ser))
    return ws


# ════════════════════════════════════════════════════════════════════════
# Entry point
# ════════════════════════════════════════════════════════════════════════

def add_figure_sheets(wb, data):
    """Add paper-figure sheets (live-formula, referencing the IPC sheet)."""
    added = []
    if IPC_SHEET not in wb.sheetnames:
        return added
    scan = _scan_ipc_speedup(wb[IPC_SHEET])
    if not scan or scan.get("mars") is None:
        return added
    cores_present = sorted({c for (c, r, p) in scan["row_of"].keys()})
    for builder in (build_F1, build_F2):
        ws = builder(wb, scan, cores_present)
        if ws is not None:
            added.append(ws.title)
    if added:
        # force Excel/LibreOffice to recompute the =IPC! formula cells on open
        # so chart series populate without a manual recalc.
        try:
            wb.calculation.fullCalcOnLoad = True
        except AttributeError:
            from openpyxl.workbook.properties import CalcProperties
            wb.calculation = CalcProperties(fullCalcOnLoad=True)
    return added
